import openpyxl
import logging
from datetime import datetime
from django.core.files import File
from user.models import Profile
from fpdf import FPDF, HTMLMixin
from excelExtract.models import (document,excel,pdfFile,
								accountEmail,excelAccount)

debug = logging.getLogger("read_excel")

class PDF(FPDF, HTMLMixin):
	# FPDF("L", "mm", "A4")
	pass

def importDataExcel(path, user=None):
	wb  	=	openpyxl.load_workbook(path,data_only=True)
	file	=	document(document=path)
	if user != None:
		file.upload_by = user 
	file.save()
		
	def sheetProcess(file,ws):
		categoryCol =  ws['BQ']
		groupCol = ws['A']

		# Get row numbers where the value is "Group"
		matching_row = 0
		for i, cell in enumerate(groupCol):
			if cell.value == "Group":
				matching_row = i + 1
				break  # Break the loop once a match is found
			if i >= 10 :
				break
		# Detect lines of data
		matching_row = matching_row or 4
		count = sum(1 for cell in groupCol[matching_row:] if cell.value is not None) + matching_row
		

		for i,row in enumerate(ws.iter_rows(min_row=matching_row + 1,max_row=count,max_col=70,values_only=True),start=matching_row):
		
			standard= str(row[1]).replace(" ", "").lower()
			acc = excelAccount.objects.filter(standardName=standard)
			if row[68] is None:
				continue
			if acc:
				try:
					excel.objects.create(filename=file,
										group=row[0],account=acc[0],postStartDate=row[4],
										postEndDate=row[5],product=row[10],
										mechanicsGetORDiscount=row[12],
										noiDungChuongTrinh=row[57],
										budgetRir=row[59],
										loaiCt=row[68])	
				except:
					if isinstance(row[4],str) and isinstance(row[5],str):
						excel.objects.create(filename=file,
											group=row[0],account=acc[0],product=row[10],
											mechanicsGetORDiscount=row[12],
											noiDungChuongTrinh=row[57],
											budgetRir=row[59],loaiCt=row[68])	
					elif isinstance(row[4],str):
						excel.objects.create(filename=file,
											group=row[0],account=acc[0],
											postEndDate=row[5],product=row[10],
											mechanicsGetORDiscount=row[12],
											noiDungChuongTrinh=row[57],
											budgetRir=row[59],loaiCt=row[68])	
					elif isinstance(row[5],str):
						excel.objects.create(filename=file,
											group=row[0],account=acc[0],
											postStartDate=row[4],product=row[10],
											mechanicsGetORDiscount=row[12],
											noiDungChuongTrinh=row[57],
											budgetRir=row[59],loaiCt=row[68])	
			else:

				excelAccount.objects.create(account=str(row[1]),standardName=standard)
				try:
					excel.objects.create(filename=file,
										group=row[0],account=row[1],
										postStartDate=row[4],
										postEndDate=row[5],
										product=row[10],
										mechanicsGetORDiscount=row[12],
										noiDungChuongTrinh=row[57],
										budgetRir=row[59],loaiCt=row[68])
				except:	
					if isinstance(row[4],str) and isinstance(row[5],str):
						excel.objects.create(filename=file,
											group=row[0],account=row[1],product=row[10],
											mechanicsGetORDiscount=row[12],
											noiDungChuongTrinh=row[57],
											budgetRir=row[59],loaiCt=row[68])	
					elif isinstance(row[4],str):
						excel.objects.create(filename=file,
											group=row[0],account=row[1],postEndDate=row[5],
											product=row[10],mechanicsGetORDiscount=row[12],
											noiDungChuongTrinh=row[57],budgetRir=row[59],loaiCt=row[68])	
					elif isinstance(row[5],str):
						excel.objects.create(filename=file,
											group=row[0],account=row[1],postStartDate=row[4],
											product=row[10],mechanicsGetORDiscount=row[12],
											noiDungChuongTrinh=row[57],budgetRir=row[59],loaiCt=row[68])	

	for sheet in wb.sheetnames:
		if sheet.strip() == "Promotion Plan BCC":
			wb.active = wb[sheet]
			ws = wb.active
			sheetProcess(file,ws)
		if sheet.strip() == "Promotion Plan FEM":
			wb.active = wb[sheet]
			ws = wb.active
			sheetProcess(file,ws)	
			

def exportFiles(fileId,user,selectedCatergoryOptions,selectedAccountOptions):
	annouceExist=""
	file = document.objects.get(pk=int(fileId))
	categoryList = list(excel.objects.filter(filename=file).values_list('loaiCt', flat=True).distinct())
	accountList = list(excel.objects.filter(filename=file).values_list('account', flat=True).distinct())
	def generateContent(categories,account):
		title = 'THÔNG BÁO VỀ CHƯƠNG TRÌNH KHUYẾN MÃI' #.encode('utf-8')
		date_year = "Tp.HCM, Ngày {}".format(str(datetime.now().date().strftime("%d/%m/%Y")))
		pdf = PDF()
		pdf.add_font('arial','',r"static/Fonts/arial.ttf",uni=True)
		pdf.add_font('arial','B',r"static/Fonts/FontsFree-Net-arial-bold.ttf",uni=True)
		pdf.add_font('arial','I',r"static/Fonts/Arial-Italic.ttf",uni=True)
		pdf.set_font('arial','', size=10)


		pdf.add_page("L")
		title_w = pdf.get_string_width(title) + 6

		doc_w = pdf.w
		center = (doc_w - title_w) / 2
		# 	Title
		pdf.text(center,15,title)
		#	Logo
		pdf.image('static/image/kimberlylogo.jpg', x = 5, w = 60,h=10, y=5,type="jpg")
		# 	Date
		pdf.set_font('arial','', size=8)
		pdf.text(255,20,date_year)
		# 	Greeting
		pdf.set_font('arial', 'B', size=9)
		text = 'Kính gửi: Quý Khách Hàng Kênh Hiện Đại'
		pdf.text(10, 25, text)
		pdf.set_font('arial','', size=8)
		text = "Công ty TNHH Kimberly-Clark Việt Nam (Công ty)  xin trân trọng thông báo chương trình đến Quý Khách Hàng như thông tin đính kèm,"

		pdf.text(10,32,text)

		pdf.set_y(pdf.get_y()+30)
		pdf.set_fill_color(153,204,255)
		pdf.cell(80,5,"Loại CT",0,0,"L",1)

		pdf.cell(80,5,",".join(str(category) for category in categories),0,1,"L",1)
		pdf.set_fill_color(153,204,255)
		pdf.set_fill_color(153,204,255)
		pdf.cell(80,5,"Account","B",0,"L",1)
		pdf.set_fill_color(153,204,255)
		pdf.cell(80,5,"{}".format(account),"B",1,"L",1)
		pdf.set_y(pdf.get_y()+5)
		# Table headers
		headers=['Mechanics: get/discount',"Product","Post start date","Post end date"]
		for i,header in enumerate(headers):
			if i==0:
				pdf.set_font('arial', 'B', size=9)
				pdf.set_fill_color(153,204,255)
				pdf.cell(80,5,"{}".format(headers[i]),"B",0,"L",1)
			elif i==1:
				pdf.set_font('arial', 'B', size=9)
				pdf.set_fill_color(153,204,255)
				pdf.cell(90,5,"{}".format(headers[i]),"B",0,"L",1)
			elif i ==2:
				pdf.set_font('arial', 'B', size=9)
				pdf.set_fill_color(153,204,255)
				pdf.cell(60,5,"{}".format(headers[i]),"B",0,"L",1)
			else:
				pdf.set_font('arial', 'B', size=9)
				pdf.set_fill_color(153,204,255)
				pdf.cell(50,5,"{}".format(headers[i]),"B",1,"L",1)

		# Table datas
		for category in categories:
			listMecha=excel.objects.filter(filename=file,account=account,loaiCt=category).values("mechanicsGetORDiscount").distinct()
			for mecha in listMecha:
				datas=excel.objects.filter(filename=file,account=account,loaiCt=category,mechanicsGetORDiscount=mecha.get("mechanicsGetORDiscount"))
				
				for row,data in enumerate(datas):
					if data.postStartDate==None or data.postEndDate==None:
						errorFlag=True
						continue
					for col,colAlphabet in enumerate(["A","B","C","D"]):
						
						mechanicsString=str(data.mechanicsGetORDiscount).replace("\n","")
						if row >0:
							mechanicsString=""	
						cellWitdhMax=80
						if pdf.get_string_width(mechanicsString) < cellWitdhMax:
							if row ==(len(datas)-1):				
								if headers[col]=='Mechanics: get/discount':
									string=mechanicsString
									pdf.cell(80,5,string,"B",0,"L")
								if headers[col]=='Product':
									string=data.product
									if string== None:
										pdf.cell(90,5,"",0,0,"L")
									else:
										pdf.cell(90,5,string,0,0,"L")
								elif headers[col]=="Post start date":
									if data.postStartDate:											
										pdf.cell(60,5,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									

								elif headers[col]=="Post end date":
									if data.postEndDate:
										pdf.cell(50,5,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
									
							else:
								if headers[col]=='Mechanics: get/discount':
									string=mechanicsString
									pdf.cell(80,5,string,0,0,"L")
								if headers[col]=='Product':
									string=data.product
									if string== None:
										pdf.cell(90,5,"",0,0,"L")
									else:
										pdf.cell(90,5,string,0,0,"L")
								elif headers[col]=="Post start date":
									if data.postStartDate:
										pdf.cell(60,5,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									
								elif headers[col]=="Post end date":
									if data.postEndDate:
										pdf.cell(50,5,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
									
						else:
							if row ==(len(datas)-1):
								mechanicsStringLen=len(mechanicsString)
								startChar=0
								maxChar=0
								holdStringEachLine=[]
								holdStringTemp=""
								while (startChar<=mechanicsStringLen):
									while (pdf.get_string_width(holdStringTemp) < cellWitdhMax) and ((startChar+maxChar)<=mechanicsStringLen):
										maxChar=maxChar+1
										holdStringTemp= mechanicsString[startChar:(maxChar+startChar)]						
									startChar=startChar+maxChar
									holdStringEachLine.append(holdStringTemp)
									#reset
									maxChar=0
									holdStringTemp=""					
							
								line=len(holdStringEachLine) #Numbers of line
								if headers[col]=='Mechanics: get/discount':
									
									xPos=pdf.get_x()
									yPos=pdf.get_y()
									pdf.multi_cell(cellWitdhMax,5,mechanicsString,"B")
									pdf.set_xy(xPos+cellWitdhMax,yPos)
								if headers[col]=='Product':
									string=data.product
									if string== None:
										pdf.cell(90,5*line,"",0,0,"L")
									else:
										pdf.cell(90,5*line,string,0,0,"L",)
								elif headers[col]=="Post start date":
									if data.postStartDate:
										pdf.cell(60,5*line,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									
								elif headers[col]=="Post end date":
									if data.postEndDate:
										pdf.cell(50,5*line,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
									
							else:
								mechanicsStringLen=len(mechanicsString)
								startChar=0
								maxChar=0
								holdStringEachLine=[]
								holdStringTemp=""
								while (startChar<=mechanicsStringLen):
									while (pdf.get_string_width(holdStringTemp) < cellWitdhMax) and ((startChar+maxChar)<=mechanicsStringLen):
										maxChar=maxChar+1
										holdStringTemp= mechanicsString[startChar:(maxChar+startChar)]						
									startChar=startChar+maxChar
									holdStringEachLine.append(holdStringTemp)
									#reset
									maxChar=0
									holdStringTemp=""					
					
								line=len(holdStringEachLine) #Numbers of line
								if headers[col]=='Mechanics: get/discount':
									
									xPos=pdf.get_x()
									yPos=pdf.get_y()
									pdf.multi_cell(cellWitdhMax,5,mechanicsString,0)
									pdf.set_xy(xPos+cellWitdhMax,yPos)
								if headers[col]=='Product':
									string=data.product
									if string== None:
										pdf.cell(90,5*line,"",0,0,"L")
									else:
										pdf.cell(90,5*line,string,0,0,"L",)
								elif headers[col]=="Post start date":
									if data.postEndDate:
										pdf.cell(60,5*line,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									
								elif headers[col]=="Post end date":
									if data.postEndDate:
										pdf.cell(50,5*line,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
								

		# End of Table
		for i,header in enumerate(headers):
			if i==0:
				pdf.set_font('arial', 'B', size=9)
				pdf.set_fill_color(153,204,255)
				pdf.cell(80,5,"","B",0,"L",1)
			elif i==1:
				pdf.set_font('arial', 'B', size=9)
				pdf.set_fill_color(153,204,255)
				pdf.cell(90,5,"","B",0,"L",1)
			elif i ==2:
				pdf.set_font('arial', 'B', size=9)
				pdf.set_fill_color(153,204,255)
				pdf.cell(60,5,"","B",0,"L",1)
			else:
				pdf.set_font('arial', 'B', size=9)
				pdf.set_fill_color(153,204,255)
				pdf.cell(50,5,"","B",1,"L",1)
				
		#	Footer
		pdf.set_y(pdf.get_y()+5)
		pdf.cell(100,5,"Mong Quý Khách Hàng cùng hợp tác với Công ty để đảm bảo các chương trình thực thi hiệu quả trong thời gian tới.",0,1)
		pdf.cell(100,5,"Nếu Quý Khách Hàng có bất kỳ vấn đề nào cần làm rõ, vui lòng cho KCV được biết để cùng trao đổi",0,1)

		pdf.set_y(pdf.get_y()+5)
		pdf.cell(40,5,"Trân trọng cảm ơn Quý Khách Hàng",0,1)
		pdf.cell(40,5,"Trưởng bộ phận quản lý kênh hiện đại",0,1)
		
		pdf.set_y(pdf.get_y()+30)
		pdf.cell(40,5,"Phạm Nguyên Thủ",0,1)
		return pdf

	if selectedCatergoryOptions == "all" and selectedAccountOptions == "all":
		print(f"{'Account':<15} {'CurrentAccountCategories'}")
		for account in accountList:
			errorFlag=False
			pdfAccountInfo = excelAccount.objects.get(account=account)
			pdfEmailDetails = accountEmail.objects.filter(account=pdfAccountInfo)
			if not pdfEmailDetails.exists():
				pdfEmailDetails = []
			# Jump to next loop
			if not any(excel.objects.filter(filename=file, account=account, loaiCt=category).exists() for category in categoryList):
				annouceExist=annouceExist+" {}_no_Data ".format(str(account))
				continue  
			currentAccountCategories = list(excel.objects.filter(filename=file,account=account).values_list('loaiCt', flat=True).distinct())
			print(f"{account:<15} {currentAccountCategories}")

			pdf = generateContent(currentAccountCategories,account)
			maxpageIndex=pdf.page-1
			pos="{}x{}".format(round(pdf.get_x()),round(pdf.get_y()))
			pdf.output("pdffile.pdf")
	
			annouceExist=annouceExist  +" {}_Success ".format(str(account))
			filename="THUTHONGBAO_{}_{}.pdf".format(str(account).upper(),str(datetime.now().date()))
			with open("pdffile.pdf",'rb') as pdf:
			
				pdffile=pdfFile() 
				if errorFlag:
					pdffile.errorFlags = True
				pdffile.masterFile = file
				pdffile.creator = user
				pdffile.account = pdfAccountInfo
				pdffile.loaict = ",".join(str(category) for category in currentAccountCategories)
				pdffile.pos = pos
				pdffile.page_number = maxpageIndex
				pdffile.slaveFile.save(filename,File(pdf))
				if pdfEmailDetails:			
					for email in pdfEmailDetails:
						pdffile.emailExtracted.add(email)
				pdffile.save()			
	
	if  selectedAccountOptions == "all" and selectedCatergoryOptions != "all":
		selectedCatergoryOptions = selectedCatergoryOptions.split(',')
		print(f"{'Account':<15} {'selectedCatergoryOptions'}")
		for account in accountList:
			errorFlag = False
			pdfAccountInfo = excelAccount.objects.get(account=account)
			pdfEmailDetails = accountEmail.objects.filter(account=pdfAccountInfo)
			if not pdfEmailDetails.exists():
				pdfEmailDetails = []
			# Jump to next loop
			if not any(excel.objects.filter(filename=file, account=account, loaiCt=category).exists() for category in selectedCatergoryOptions):
				annouceExist=annouceExist+" {}_no_Data ".format(str(account))
				continue  
			print(f"{account:<15} {selectedCatergoryOptions}")
			

			pdf = generateContent(selectedAccountOptions,account)
			pos = "{}x{}".format(round(pdf.get_x()),round(pdf.get_y()))
			maxpageIndex = pdf.page-1
			pdf.output("pdffile.pdf")
			annouceExist=annouceExist  +" {}_Success ".format(str(account))
			filename="THUTHONGBAO_{}_{}.pdf".format(str(account).upper(),str(datetime.now().date()))
			
			with open("pdffile.pdf",'rb') as pdf:
			
				pdffile=pdfFile()
				pdffile.masterFile=file
				if errorFlag:
					pdffile.errorFlags=True
				pdffile.creator=user
				pdffile.account=pdfAccountInfo
				pdffile.loaict=",".join(str(category) for category in selectedCatergoryOptions)
				pdffile.pos=pos
				pdffile.page_number=maxpageIndex
				pdffile.slaveFile.save(filename,File(pdf))
				if pdfEmailDetails:			
					for email in pdfEmailDetails:
						pdffile.emailExtracted.add(email)
				pdffile.save()
	if  selectedAccountOptions != "all" and selectedCatergoryOptions == "all":
		selectedAccountOptions = selectedCatergoryOptions.split(',')
		print(f"{'Account':<15} {'Catergories'}")
		for account in selectedAccountOptions:
			errorFlag = False
			pdfAccountInfo = excelAccount.objects.get(account=account)
			pdfEmailDetails = accountEmail.objects.filter(account=pdfAccountInfo)
			if not pdfEmailDetails.exists():
				pdfEmailDetails = []
			# Jump to next loop
			if not any(excel.objects.filter(filename=file, account=account, loaiCt=category).exists() for category in categoryList):
				annouceExist=annouceExist+" {}_no_Data ".format(str(account))
				continue  
			currentAccountCategories = list(excel.objects.filter(filename=file,account=account).values_list('loaiCt', flat=True).distinct())
			print(f"{account:<15} {currentAccountCategories}")

			pdf = generateContent(currentAccountCategories,account)
			maxpageIndex=pdf.page-1
			pos="{}x{}".format(round(pdf.get_x()),round(pdf.get_y()))
			pdf.output("pdffile.pdf")
			annouceExist=annouceExist  +" {}_Success ".format(str(account))
			filename="THUTHONGBAO_{}_{}.pdf".format(str(account).upper(),str(datetime.now().date()))
			with open("pdffile.pdf",'rb') as pdf:
			
				pdffile=pdfFile()
				if errorFlag:
					pdffile.errorFlags = True
				pdffile.masterFile = file
				pdffile.creator = user
				pdffile.account = pdfAccountInfo
				pdffile.loaict = ",".join(str(category) for category in currentAccountCategories)
				pdffile.pos = pos
				pdffile.page_number = maxpageIndex
				pdffile.slaveFile.save(filename,File(pdf))
				if pdfEmailDetails:			
					for email in pdfEmailDetails:
						pdffile.emailExtracted.add(email)
				pdffile.save()			
	
	if  selectedAccountOptions != "all" and selectedCatergoryOptions != "all":
		selectedAccountOptions = selectedAccountOptions.split(',')
		selectedCatergoryOptions = selectedCatergoryOptions.split(',')
		print(selectedAccountOptions , selectedCatergoryOptions)
		print(f"{'Account':<15} {'Catergories'}")
		for account in selectedAccountOptions:
			errorFlag = False
			pdfAccountInfo = excelAccount.objects.get(account=account)
			pdfEmailDetails = accountEmail.objects.filter(account=pdfAccountInfo)
			if not pdfEmailDetails.exists():
				pdfEmailDetails = []
			# Jump to next loop
			if not any(excel.objects.filter(filename=file, account=account, loaiCt=category).exists() for category in selectedCatergoryOptions):
				annouceExist=annouceExist+" {}_no_Data ".format(str(account))
				continue  
			print(f"{account:<15} {selectedCatergoryOptions}")
			pdf = generateContent(selectedCatergoryOptions,account)
			pos = "{}x{}".format(round(pdf.get_x()),round(pdf.get_y()))
			maxpageIndex = pdf.page-1
			pdf.output("pdffile.pdf")
			annouceExist=annouceExist  +" {}_Success ".format(str(account))
			filename="THUTHONGBAO_{}_{}.pdf".format(str(account).upper(),str(datetime.now().date()))
			
			with open("pdffile.pdf",'rb') as pdf:
			
				pdffile=pdfFile()
				pdffile.masterFile=file
				if errorFlag:
					pdffile.errorFlags=True
				pdffile.creator=user
				pdffile.account=pdfAccountInfo
				pdffile.loaict=",".join(str(category) for category in selectedCatergoryOptions)
				pdffile.pos=pos
				pdffile.page_number=maxpageIndex
				pdffile.slaveFile.save(filename,File(pdf))
				if pdfEmailDetails:			
					for email in pdfEmailDetails:
						pdffile.emailExtracted.add(email)
				pdffile.save()
	return annouceExist


