from datetime import datetime
from .models import pdfFile
import openpyxl
from datetime import datetime
from django.template.loader import get_template
from django.core.mail import EmailMessage
from django.conf import settings
from openpyxl.styles import Alignment,NamedStyle,PatternFill,Font
from io import BytesIO
import openpyxl
import requests, logging
logger_debug = logging.getLogger("auto_email")
def is_leap_year(year): 
	if year % 100 == 0:
		return year % 100 == 0

	return year % 4 == 0

def get_lapse():
	last_month = datetime.today().month - 1
	current_year = datetime.today().year

	#is last month a month with 30 days?
	if last_month in [9, 4, 6, 11]:
		lapse = 30

	#is last month a month with 31 days?
	elif last_month in [1, 3, 5, 7, 8, 10, 12]:
		lapse = 31

	#is last month February?
	else:
		if is_leap_year(current_year):
			lapse = 29
		else:
			lapse = 30

	return lapse
def get_date_range(lapse):
	year= datetime.today().year
	last_month = datetime.today().month - 1
	first_date = datetime(year, last_month, 1)

	end_date = datetime(year, last_month, int(lapse))
	first_date_new_month=datetime(year, datetime.today().month, 1)
	return first_date,end_date,first_date_new_month
def auto_report_excel(from_date,to_date):
	col_names = ["","Account","Category","File","CreatedTime","ConfirmedTime","SendedTime","Creator","Confirmer","Sender","Confirmed","Signed","Sended"]
	actlogs = pdfFile.objects.filter(SignedTime__date__gte=from_date,SignedTime__date__lte=to_date,signed=True)

	wb = openpyxl.Workbook()
	wb.iso_dates = True
	ws = wb['Sheet']
	ws.title='Report'


	# create title
	light_blue_fill = PatternFill(start_color='87CEFA',
                   end_color='87CEFA',
                   fill_type='solid')
	font = Font(name='Arial', size=12, bold=False,
              	vertAlign=None, underline='none', strike=False,
                color='FF000000')
	for col in range(1, len(col_names)):
		_ = ws.cell(column=col,row=1,value=col_names[col])
		_.fill = light_blue_fill
		_.font = font
	alignment = Alignment(horizontal='general')
	normal_format = NamedStyle(name="normal",alignment=alignment)
	datetime_format = NamedStyle(name="datetime",number_format="DD/MMM/YYYY h:mm",alignment=alignment)
	date_format = NamedStyle(name="date",number_format="DD/MMM/YYYY",alignment=alignment)
	ws.column_dimensions['A'].width = 20
	ws.column_dimensions['B'].width = 40
	ws.column_dimensions['C'].width = 40
	ws.column_dimensions['D'].width = 20
	ws.column_dimensions['E'].width = 20
	ws.column_dimensions['F'].width = 20
	ws.column_dimensions['G'].width = 20
	ws.column_dimensions['H'].width = 20
	ws.column_dimensions['l'].width = 20

	# create content
	# now = pytz.utc.localize(datetime.utcnow())
	# now = now.replace(tzinfo=None)
	# print(now)
	for row, ticket in enumerate(actlogs, start=2):
		for col in range(1,13):		
			if col_names[col] == "Account" :
				if ticket.account:
					c = ws.cell(column=col,row=row,value=str(ticket.account))
			if col_names[col] == "Category" :
				if ticket.loaict:
					c = ws.cell(column=col,row=row,value=str(ticket.loaict))
			if col_names[col] == "file" :
				if ticket.slaveFile:
					c = ws.cell(column=col,row=row,value=str(ticket.slaveFile).replace("documents/slavefiles/","") )
			if col_names[col] == "file" :
				if ticket.slaveFile:
					c = ws.cell(column=col,row=row,value=str(ticket.slaveFile).replace("documents/slavefiles/","") )
				else:
					c = ws.cell(column=col,row=row,value="")
				c.style = datetime_format
				
			elif col_names[col] == "creator":
				c = ws.cell(column=col,row=row,value=str(ticket.creator))
				c.style = date_format
			elif col_names[col] == "CreatedTime":
				if ticket.createdTime:
					c = ws.cell(column=col,row=row,value=str(ticket.createdTime.date()))
				else:
					c = ws.cell(column=col,row=row,value="")
				c.style = normal_format
			elif col_names[col] == "ConfirmedTime":
				if ticket.confirmedTime:
					c = ws.cell(column=col,row=row,value=str(ticket.confirmedTime.date()))
				else:
					c = ws.cell(column=col,row=row,value="")
				c.style = normal_format
			elif col_names[col] == "SendedTime":
				if ticket.sendingTime:
					c = ws.cell(column=col,row=row,value=str(ticket.sendingTime.date()))
					c.style = normal_format
				else:
					c = ws.cell(column=col,row=row,value="")
			elif col_names[col] == "creator":
				if ticket.sendingTime:
					c = ws.cell(column=col,row=row,value=str(ticket.creator))
					c.style = normal_format
			elif col_names[col] == "confirmer":
				if ticket.confirmer:
					c = ws.cell(column=col,row=row,value=str(ticket.confirmer))
					c.style = normal_format
			elif col_names[col] == "sender":
				if ticket.signer:
					c = ws.cell(column=col,row=row,value=str(ticket.signer))
					c.style = normal_format		
			elif col_names[col] == "sended":
				if ticket.sended:
					c = ws.cell(column=col,row=row,value=str(ticket.sended))
					c.style = normal_format		
			elif col_names[col] == "Sended":
				if ticket.sended:
					c = ws.cell(column=col,row=row,value=str(ticket.sended))
					c.style = normal_format
			elif col_names[col] == "Signed":
				if ticket.signed:
					c = ws.cell(column=col,row=row,value=str(ticket.signed))
					c.style = normal_format	
			elif col_names[col] == "Confirmed":
				if ticket.confirmed:
					c = ws.cell(column=col,row=row,value=str(ticket.confirmed))
					c.style = normal_format		
	return wb
def send_report():
	last_month = datetime.today().month - 1
	subject = "Report File hàng tháng - Psign tháng {}".format(last_month)
	date_range=get_date_range(get_lapse())
	len_query = len(pdfFile.objects.filter(SignedTime__date__gte=date_range[0].strftime("%Y-%m-%d"),SignedTime__date__lte=date_range[2].strftime("%Y-%m-%d"),signed=True))
	html_message = get_template("template_email_summary_report.html").render({"first_date":date_range[0].strftime("%d/%m/%Y"),"end_date":date_range[1].strftime("%d/%m/%Y"),"file_nums":len_query})

	msg = EmailMessage(subject,html_message,settings.EMAIL_HOST_USER,to=['longnld@pvs.com.vn','dk@pvs.com.vn'])
	msg.content_subtype = "html"
	output=BytesIO()
	auto_report_excel(date_range[0].strftime("%Y-%m-%d"),date_range[2].strftime("%Y-%m-%d")).save(output)
	msg.attach('{}_{}.xlsx'.format(date_range[0].strftime("%Y-%m-%d"),date_range[1].strftime("%Y-%m-%d")),output.getvalue(),'application/vnd.ms-excel')
		
	msg.send()