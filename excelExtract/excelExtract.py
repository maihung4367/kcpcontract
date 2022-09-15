from urllib import request
import openpyxl
from datetime import datetime
from excelExtract.models import document,excel,pdfFile,accountEmail,excelAccount
import os
from django.core.files import File
# import json
# import pdfkit 
from user.models import Profile
from fpdf import FPDF, HTMLMixin
from datetime import datetime
import re
class PDF(FPDF, HTMLMixin):
	# FPDF("L", "mm", "A4")
	
	pass
def importDataExcel(path, user=None):
	wb=openpyxl.load_workbook(path,data_only=True)
	file=document(document=path)
	if user != None:
		file.upload_by = user
	file.save()
	
	#TÌM HEADERS,LOẠI CT TRONG SHEET CẦN TÌM/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
	programCate=[]	#LIST TỔNG HỢP LẠI CHƯƠNG TRÌNH CỦA 2 BIẾN DƯỚI (ĐỀ PHÒNG TRƯỜNG HỢP CÓ LOẠI CT MỚI MÀ SHEET KHÁC KHÔNG CÓ)
	programCate1=[] #CHỨA CÁC LOẠI CT TRONG Ecom Promotion Plan BCC_for SO
	programCate2=[] #CHỨA CÁC LOẠI CT TRONG  MnB Promotion Plan BCC_for SO
	listHeader=[] 	#CHỨA HEADER của Ecom VÀ Mnb Promotion Plan BCC_for SO (HIỆN TẠI DỰA VÀO ECOM ĐỂ XÁC ĐỊNH VÌ CẢ 2 GIỐNG NHAU )
	
	for f in wb.sheetnames:
		if f=="Ecom Promotion Plan BCC_for SO":
			wb.active=wb[f]
			ws=wb.active			
			programCateCol=ws['BQ'] # CỘT LOẠI CT
			group=ws['A'] # CỘT GROUP
			lineEnd=0 # XÁC ĐỊNH ROWS DỮ LIỆU  KẾT THÚC TẠI DÒNG NÀO ,DỰA TRÊN GROUP
			count=0 # ĐẾM SỐ LƯỢNG DATA IMPORT
			#LIỆT KÊ LOẠI CT
			for i in range(3,len(programCateCol)): 
				if  programCateCol[i].value not in programCate1 :
					programCate1.append(programCateCol[i].value)
					if  programCateCol[i].value==None:
						programCate1.pop()
						break
			# XÁC ĐỊNH ROWS DỮ LIỆU  KẾT THÚC TẠI DÒNG NÀO ,DỰA TRÊN GROUP
			for i in range(3,len(group)):		
				if group[i].value != None:
					count=count+1
			lineEnd=count+3
			print("lineEnd" +"{}: {}".format(str(f),lineEnd))
			#import data
			rangeline=lineEnd-1
			for i,row in enumerate(ws.rows):
				
				if i>=3 and i<=rangeline:
					excel.objects.create(filename=file,group=row[0].value,account=row[1].value,postStartDate=row[4].value,postEndDate=row[5].value,product=row[10].value,mechanicsGetORDiscount=row[12].value,noiDungChuongTrinh=row[57].value,budgetRir=row[59].value,loaiCt=row[68].value)
					try:				
						acc=excelAccount.objects.filter(account=row[1].value)
						print(acc)
						if not acc :
							print("1")
							newAccount=excelAccount.objects.create(account=row[1].value)
							newAccount.save()
							if row[70].value !=None:
								email=accountEmail.objects.create(email=row[70].value)	
								email.save()
						else:
							print("2")
							emailfilter=accountEmail.objects.filter(account=acc[0],email=row[70].value)
							if not emailfilter:
								if row[70].value != None:
									email=accountEmail.objects.create(account=acc[0],email=row[70].value)
									email.save()
					except:
						pass
					
				elif i> lineEnd:
					break
		if f=="MnB Promotion Plan BCC_for SO":
			wb.active=wb[f]
			ws=wb.active	
			programCateCol=ws['BQ'] # CỘT LOẠI CT
			group=ws['A']# CỘT GROUP
			count=0 # ĐẾM SỐ LƯỢNG DATA IMPORT
			lineEnd=0
			#LIỆT KÊ LOẠI CT
			for i in range(3,len(programCateCol)): 
				if  programCateCol[i].value not in programCate2 :
					programCate2.append(programCateCol[i].value)
					if  programCateCol[i].value==None:
						programCate2.pop()
						break
			# XÁC ĐỊNH ROWS DỮ LIỆU  KẾT THÚC TẠI DÒNG NÀO ,DỰA TRÊN GROUP
			for i in range(3,len(group)):
				if group[i].value != None:
					count=count+1
			lineEnd=count+3
			print("lineEnd" +"{}: {}".format(str(f),lineEnd))
			
			# for i,row in enumerate(ws.rows,start=4):
			# 	while i <= lineEnd:
			# 		print(row[3].value)
			#import data
			rangeline=lineEnd-1
			for i,row in enumerate(ws.rows):
				if i>=3 and i<=rangeline:
					excel.objects.create(filename=file,group=row[0].value,account=row[1].value,postStartDate=row[4].value,postEndDate=row[5].value,product=row[10].value,mechanicsGetORDiscount=row[12].value,noiDungChuongTrinh=row[57].value,budgetRir=row[59].value,loaiCt=row[68].value)
					try:				
						acc=excelAccount.objects.filter(account=row[1].value)
						print(acc)
						if not acc :
							print("1")
							newAccount=excelAccount.objects.create(account=row[1].value)
							newAccount.save()
							if row[70].value !=None:
								email=accountEmail.objects.create(email=row[70].value)	
								email.save()
						else:
							print("2")
							emailfilter=accountEmail.objects.filter(account=acc[0],email=row[70].value)
							if not emailfilter:
								if row[70].value != None:
									email=accountEmail.objects.create(account=acc[0],email=row[70].value)
									email.save()
					except:
						pass
				elif i> lineEnd:
					break
			
	# TỔNG HỢP LẠI LOẠI CHƯƠNG TRÌNH CỦA 2 LIST (ĐỀ PHÒNG TRƯỜNG HỢP CÓ LOẠI CT MỚI MÀ LISTS KHÁC KHÔNG CÓ)
	for i in programCate1:
		if i not in programCate:
			programCate.append(i)
	for i in programCate2:
		if i not in programCate:
			programCate.append(i)
	return listHeader
	#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////

# excelExtract.exportFiles(loaict,fileID,loaiAccount)
def exportFiles(loaict,fileID,loaiAccount,user):
	print("fileID :{}".format(fileID))
	print("loaict :{}".format(loaict))
	print("loaiAccount :{}".format(loaiAccount))
	print(type(loaiAccount))
	annouceExist=""
	#XÁC ĐỊNH LOẠI CT  FILE
	listloaict=[]
	listAccount=[]
	file=document.objects.get(pk=int(fileID))
	print(file)
	for f in excel.objects.filter(filename=file):
		if f.loaiCt not in listloaict:
			listloaict.append(f.loaiCt)
		if f.account not in listAccount:
			listAccount.append(f.account)	
	print(listloaict)
	print(listAccount)
	# print(listNoiDungCt)
	if loaiAccount=="all" and loaict=="all": 
		
		for f in listAccount:
			emptyList=[]
			
			pdfAccount=excelAccount.objects.get(account=f)
			try:
				pdfEmail=accountEmail.objects.filter(account=pdfAccount)
			except:
				pdfEmail=[]
			for chuongTrinh in listloaict:
				if excel.objects.filter(filename=file,account=f,loaiCt=chuongTrinh).exists():
					emptyList=[]
					continue
				else:
					emptyList.append("{}".format(str(chuongTrinh)))

			if emptyList != []:
				annouceExist=annouceExist+" {}_no_Data ".format(str(f))
				continue
			else:
				print("111111111111111111111111111")
				annouceExist=annouceExist+" {}_Success ".format(str(f))
			title = 'THÔNG BÁO VỀ CHƯƠNG TRÌNH KHUYẾN MÃI' #.encode('utf-8')
			date_year = "Tp.HCM, Ngày {}".format(str(datetime.now().date().strftime("%d/%m/%Y")))
			pdf = PDF()
			pdf.add_font('arial','',r"static/Fonts/arial.ttf",uni=True)
			pdf.add_font('arial','B',r"static/Fonts/FontsFree-Net-arial-bold.ttf",uni=True)
			pdf.add_font('arial','I',r"static/Fonts/Arial-Italic.ttf",uni=True)
			pdf.set_font('arial','', size=10)
			# pdf=FPDF("L", "mm", "A4")

			pdf.add_page("L")
			title_w = pdf.get_string_width(title) + 6

			doc_w = pdf.w
			center = (doc_w - title_w) / 2
			#tittle
			pdf.text(center,15,title)
			#logo
			pdf.image('static/image/kimberlylogo.jpg', x = 5, w = 60,h=10, y=5,type="jpg")
			# date year

			pdf.set_font('arial','', size=8)
			pdf.text(255,20,date_year)
			# Kính gửi
			pdf.set_font('arial', 'B', size=9)
			text = 'Kính gửi: Quý Khách Hàng Kênh Hiện Đại'

			pdf.text(10, 25, text)
			# Tên chương trình
			pdf.set_font('arial','', size=8)
			text = "Công ty TNHH Kimberly-Clark Việt Nam (Công ty)  xin trân trọng thông báo chương trình đến Quý Khách Hàng như thông tin đính kèm,"

			pdf.text(10,32,text)
			# Loại CT

			#loai ct,loai acc
			print(pdf.get_x())
			print(pdf.get_y())
			pdf.set_y(pdf.get_y()+30)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"Loại CT",0,0,"L",1)

			pdf.cell(80,5,"{}".format(",".join(listloaict)),0,1,"L",1)
			pdf.set_fill_color(153,204,255)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"Account","B",0,"L",1)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"{}".format(f),"B",1,"L",1)
			pdf.set_y(pdf.get_y()+5)
			#table header
			headers=['Mechanics: get/discount',"Product","Post start date","Post end date"]
			for i,header in enumerate(headers):
				if i==0:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"{}".format(headers[i]),"B",0,"L",1)
				elif i==1:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"{}".format(headers[i]),"B",0,"L",1)
				elif i ==2:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"{}".format(headers[i]),"B",0,"L",1)
				else:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"{}".format(headers[i]),"B",1,"L",1)

			#table data
			for ct in listloaict:
				listMecha=excel.objects.filter(filename=file,account=f,loaiCt=ct).values("mechanicsGetORDiscount").distinct()
				for mecha in listMecha:
					datas=excel.objects.filter(filename=file,account=f,loaiCt=ct,mechanicsGetORDiscount=mecha.get("mechanicsGetORDiscount"))
					print(datas)
					for row,data in enumerate(datas):
						for col,colAlphabet in enumerate(["A","B","C","D"]):
							
							mechanicsString=data.mechanicsGetORDiscount.replace("\n","")
							if row >0:
								mechanicsString=""	
							cellWitdhMax=80
							if pdf.get_string_width(mechanicsString) < cellWitdhMax:
								if row ==(len(datas)-1):				
									if headers[col]=='Mechanics: get/discount':
										string=mechanicsString
										pdf.cell(80,5,string,"B",0,"L")
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(80,5,"",0,0,"L")
										else:
											pdf.cell(80,5,string,0,0,"L")
									elif headers[col]=="Post start date":
										pdf.cell(60,5,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									elif headers[col]=="Post end date":
										pdf.cell(60,5,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
								else:
									if headers[col]=='Mechanics: get/discount':
										string=mechanicsString
										pdf.cell(80,5,string,0,0,"L")
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(80,5,"",0,0,"L")
										else:
											pdf.cell(80,5,string,0,0,"L")
									elif headers[col]=="Post start date":
										pdf.cell(60,5,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									elif headers[col]=="Post end date":
										pdf.cell(60,5,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
							else:
								if row ==(len(datas)-1):
									mechanicsStringLen=len(mechanicsString)
									startChar=0
									maxChar=0
									holdStringEachLine=[]
									holdStringTemp=""
									while (startChar<=mechanicsStringLen):
										while (pdf.get_string_width(holdStringTemp) < cellWitdhMax) and ((startChar+maxChar)<=mechanicsStringLen):
											maxChar=maxChar+1
											holdStringTemp= mechanicsString[startChar:(maxChar+startChar)]						
										startChar=startChar+maxChar
										holdStringEachLine.append(holdStringTemp)
										#reset
										maxChar=0
										holdStringTemp=""					
									print(holdStringEachLine)
									line=len(holdStringEachLine) #Numbers of line
									if headers[col]=='Mechanics: get/discount':
										
										xPos=pdf.get_x()
										yPos=pdf.get_y()
										pdf.multi_cell(cellWitdhMax,5,mechanicsString,"B")
										pdf.set_xy(xPos+cellWitdhMax,yPos)
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(80,5*line,"",0,0,"L")
										else:
											pdf.cell(80,5*line,string,0,0,"L",)
									elif headers[col]=="Post start date":
										pdf.cell(60,5*line,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									elif headers[col]=="Post end date":
										pdf.cell(60,5*line,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
								else:
									mechanicsStringLen=len(mechanicsString)
									startChar=0
									maxChar=0
									holdStringEachLine=[]
									holdStringTemp=""
									while (startChar<=mechanicsStringLen):
										while (pdf.get_string_width(holdStringTemp) < cellWitdhMax) and ((startChar+maxChar)<=mechanicsStringLen):
											maxChar=maxChar+1
											holdStringTemp= mechanicsString[startChar:(maxChar+startChar)]						
										startChar=startChar+maxChar
										holdStringEachLine.append(holdStringTemp)
										#reset
										maxChar=0
										holdStringTemp=""					
									print(holdStringEachLine)
									line=len(holdStringEachLine) #Numbers of line
									if headers[col]=='Mechanics: get/discount':
										
										xPos=pdf.get_x()
										yPos=pdf.get_y()
										pdf.multi_cell(cellWitdhMax,5,mechanicsString,0)
										pdf.set_xy(xPos+cellWitdhMax,yPos)
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(80,5*line,"",0,0,"L")
										else:
											pdf.cell(80,5*line,string,0,0,"L",)
									elif headers[col]=="Post start date":
										pdf.cell(60,5*line,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									elif headers[col]=="Post end date":
										pdf.cell(60,5*line,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")

			#table footer
			for i,header in enumerate(headers):
				if i==0:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"","B",0,"L",1)
				elif i==1:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"","B",0,"L",1)
				elif i ==2:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"","B",0,"L",1)
				else:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"","B",1,"L",1)
					
			#add footer
			pdf.set_y(pdf.get_y()+5)
			pdf.cell(100,5,"Mong Quý Khách Hàng cùng hợp tác với Công ty để đảm bảo các chương trình thực thi hiệu quả trong thời gian tới.",0,1)
			pdf.cell(100,5,"Nếu Quý Khách Hàng có bất kỳ vấn đề nào cần làm rõ, vui lòng cho KCV được biết để cùng trao đổi",0,1)

			pdf.set_y(pdf.get_y()+5)
			pdf.cell(40,5,"Trân trọng cảm ơn Quý Khách Hàng",0,1)
			pdf.cell(40,5,"Trưởng bộ phận quản lý kênh hiện đại",0,1)
			
			
			pdf.set_y(pdf.get_y()+5)
			pdf.cell(40,5,"Phạm Nguyên Thủ",0,1)
			
			
		
			pos="{}x{}".format(round(pdf.get_x()),round(pdf.get_y()))
			print(pos)
			maxpageIndex=pdf.page-1
			print(maxpageIndex)
			pdf.output("pdffile.pdf")
			# pdf.write_html(html)
			filename="ThưThôngBáo_{}_{}.pdf".format(f,str(datetime.now().date()))
			print(filename)
			with open("pdffile.pdf",'rb') as pdf:
				print(1)
				pdffile=pdfFile()

				pdffile.masterFile=file
				pdffile.creator=user
				pdffile.account=pdfAccount
				pdffile.loaict=",".join(listloaict)
				pdffile.pos=pos
				pdffile.page_number=maxpageIndex
				pdffile.slaveFile.save(filename,File(pdf))
				if pdfEmail:			
					for email in pdfEmail:
						print(email)
						pdffile.emailExtracted.add(email)
				pdffile.save()
	elif  loaiAccount=="all" and loaict!="all":
		print("22222222222222222222222222222")
		listCt=loaict.split(',')
		for f in listAccount:
			emptyList=[]
			pdfAccount=excelAccount.objects.get(account=f)
			try:
				pdfEmail=accountEmail.objects.filter(account=pdfAccount)
			except:
				pdfEmail=[]
			for chuongTrinh in listCt:
				if excel.objects.filter(filename=file,account=f,loaiCt=chuongTrinh).exists():
					emptyList=[]
					continue
				else:
					emptyList.append("{}".format(str(chuongTrinh)))

			if emptyList != []:
				annouceExist=annouceExist+" {}_no_Data ".format(str(f))
				continue
			else:
				annouceExist=annouceExist+" {}_Success ".format(str(f))
			title = 'THÔNG BÁO VỀ CHƯƠNG TRÌNH KHUYẾN MÃI' #.encode('utf-8')
			date_year = "Tp.HCM, Ngày {}".format(str(datetime.now().date().strftime("%d/%m/%Y")))
			pdf = PDF()
			pdf.add_font('arial','',r"static/Fonts/arial.ttf",uni=True)
			pdf.add_font('arial','B',r"static/Fonts/FontsFree-Net-arial-bold.ttf",uni=True)
			pdf.add_font('arial','I',r"static/Fonts/Arial-Italic.ttf",uni=True)
			pdf.set_font('arial','', size=10)
			# pdf=FPDF("L", "mm", "A4")

			pdf.add_page("L")
			title_w = pdf.get_string_width(title) + 6

			doc_w = pdf.w
			center = (doc_w - title_w) / 2
			#tittle
			pdf.text(center,15,title)
			#logo
			pdf.image('static/image/kimberlylogo.jpg', x = 5, w = 60,h=10, y=5,type="jpg")
			# date year

			pdf.set_font('arial','', size=8)
			pdf.text(255,20,date_year)
			# Kính gửi
			pdf.set_font('arial', 'B', size=9)
			text = 'Kính gửi: Quý Khách Hàng Kênh Hiện Đại'

			pdf.text(10, 25, text)
			# Tên chương trình
			pdf.set_font('arial','', size=8)
			text = "Công ty TNHH Kimberly-Clark Việt Nam (Công ty)  xin trân trọng thông báo chương trình đến Quý Khách Hàng như thông tin đính kèm,"

			pdf.text(10,32,text)
			# Loại CT

			#loai ct,loai acc
			print(pdf.get_x())
			print(pdf.get_y())
			pdf.set_y(pdf.get_y()+30)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"Loại CT",0,0,"L",1)

			pdf.cell(80,5,"{}".format(loaict),0,1,"L",1)
			pdf.set_fill_color(153,204,255)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"Account","B",0,"L",1)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"{}".format(f),"B",1,"L",1)
			pdf.set_y(pdf.get_y()+5)
			#table header
			headers=['Mechanics: get/discount',"Product","Post start date","Post end date"]
			for i,header in enumerate(headers):
				if i==0:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"{}".format(headers[i]),"B",0,"L",1)
				elif i==1:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"{}".format(headers[i]),"B",0,"L",1)
				elif i ==2:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"{}".format(headers[i]),"B",0,"L",1)
				else:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"{}".format(headers[i]),"B",1,"L",1)

			#table data
			for ct in listCt:
				listMecha=excel.objects.filter(filename=file,account=f,loaiCt=ct).values("mechanicsGetORDiscount").distinct()
				for mecha in listMecha:
					datas=excel.objects.filter(filename=file,account=f,loaiCt=ct,mechanicsGetORDiscount=mecha.get("mechanicsGetORDiscount"))
					print(datas)
					for row,data in enumerate(datas):
						for col,colAlphabet in enumerate(["A","B","C","D"]):
							
							mechanicsString=data.mechanicsGetORDiscount.replace("\n","")
							if row >0:
								mechanicsString=""	
							cellWitdhMax=80
							if pdf.get_string_width(mechanicsString) < cellWitdhMax:
								if row ==(len(datas)-1):				
									if headers[col]=='Mechanics: get/discount':
										string=mechanicsString
										pdf.cell(80,5,string,"B",0,"L")
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(80,5,"",0,0,"L")
										else:
											pdf.cell(80,5,string,0,0,"L")
									elif headers[col]=="Post start date":
										pdf.cell(60,5,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									elif headers[col]=="Post end date":
										pdf.cell(60,5,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
								else:
									if headers[col]=='Mechanics: get/discount':
										string=mechanicsString
										pdf.cell(80,5,string,0,0,"L")
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(80,5,"",0,0,"L")
										else:
											pdf.cell(80,5,string,0,0,"L")
									elif headers[col]=="Post start date":
										pdf.cell(60,5,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									elif headers[col]=="Post end date":
										pdf.cell(60,5,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
							else:
								if row ==(len(datas)-1):
									mechanicsStringLen=len(mechanicsString)
									startChar=0
									maxChar=0
									holdStringEachLine=[]
									holdStringTemp=""
									while (startChar<=mechanicsStringLen):
										while (pdf.get_string_width(holdStringTemp) < cellWitdhMax) and ((startChar+maxChar)<=mechanicsStringLen):
											maxChar=maxChar+1
											holdStringTemp= mechanicsString[startChar:(maxChar+startChar)]						
										startChar=startChar+maxChar
										holdStringEachLine.append(holdStringTemp)
										#reset
										maxChar=0
										holdStringTemp=""					
									print(holdStringEachLine)
									line=len(holdStringEachLine) #Numbers of line
									if headers[col]=='Mechanics: get/discount':
										
										xPos=pdf.get_x()
										yPos=pdf.get_y()
										pdf.multi_cell(cellWitdhMax,5,mechanicsString,"B")
										pdf.set_xy(xPos+cellWitdhMax,yPos)
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(80,5*line,"",0,0,"L")
										else:
											pdf.cell(80,5*line,string,0,0,"L",)
									elif headers[col]=="Post start date":
										pdf.cell(60,5*line,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									elif headers[col]=="Post end date":
										pdf.cell(60,5*line,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
								else:
									mechanicsStringLen=len(mechanicsString)
									startChar=0
									maxChar=0
									holdStringEachLine=[]
									holdStringTemp=""
									while (startChar<=mechanicsStringLen):
										while (pdf.get_string_width(holdStringTemp) < cellWitdhMax) and ((startChar+maxChar)<=mechanicsStringLen):
											maxChar=maxChar+1
											holdStringTemp= mechanicsString[startChar:(maxChar+startChar)]						
										startChar=startChar+maxChar
										holdStringEachLine.append(holdStringTemp)
										#reset
										maxChar=0
										holdStringTemp=""					
									print(holdStringEachLine)
									line=len(holdStringEachLine) #Numbers of line
									if headers[col]=='Mechanics: get/discount':
										
										xPos=pdf.get_x()
										yPos=pdf.get_y()
										pdf.multi_cell(cellWitdhMax,5,mechanicsString,0)
										pdf.set_xy(xPos+cellWitdhMax,yPos)
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(80,5*line,"",0,0,"L")
										else:
											pdf.cell(80,5*line,string,0,0,"L",)
									elif headers[col]=="Post start date":
										pdf.cell(60,5*line,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									elif headers[col]=="Post end date":
										pdf.cell(60,5*line,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")

			#table footer
			for i,header in enumerate(headers):
				if i==0:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"","B",0,"L",1)
				elif i==1:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"","B",0,"L",1)
				elif i ==2:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"","B",0,"L",1)
				else:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"","B",1,"L",1)
					
			#add footer
			pdf.set_y(pdf.get_y()+5)
			pdf.cell(100,5,"Mong Quý Khách Hàng cùng hợp tác với Công ty để đảm bảo các chương trình thực thi hiệu quả trong thời gian tới.",0,1)
			pdf.cell(100,5,"Nếu Quý Khách Hàng có bất kỳ vấn đề nào cần làm rõ, vui lòng cho KCV được biết để cùng trao đổi",0,1)

			pdf.set_y(pdf.get_y()+5)
			pdf.cell(40,5,"Trân trọng cảm ơn Quý Khách Hàng",0,1)
			pdf.cell(40,5,"Trưởng bộ phận quản lý kênh hiện đại",0,1)
			
			
			pdf.set_y(pdf.get_y()+5)
			pdf.cell(40,5,"Phạm Nguyên Thủ",0,1)
			
			
		
			pos="{}x{}".format(round(pdf.get_x()),round(pdf.get_y()))
			print(pos)
			maxpageIndex=pdf.page-1
			print(maxpageIndex)
			pdf.output("pdffile.pdf")
			# pdf.write_html(html)
			filename="ThưThôngBáo_{}_{}.pdf".format(f,str(datetime.now().date()))
			print(filename)
			with open("pdffile.pdf",'rb') as pdf:
				print(1)
				pdffile=pdfFile()

				pdffile.masterFile=file
				pdffile.creator=user
				pdffile.account=pdfAccount
				pdffile.loaict=loaict
				pdffile.pos=pos
				pdffile.page_number=maxpageIndex
				pdffile.slaveFile.save(filename,File(pdf))
				if pdfEmail:			
					for email in pdfEmail:
						print(email)
						pdffile.emailExtracted.add(email)
				pdffile.save()
	elif loaiAccount!="all" and loaict=="all":
		print("#33333333333333333333333")
		listAcc=loaiAccount.split(',')	
		print(listAcc)
		for f in listAcc:
			emptyList=[]
			pdfAccount=excelAccount.objects.get(account=f)
			try:
				pdfEmail=accountEmail.objects.filter(account=pdfAccount)
			except:
				pdfEmail=[]
			for chuongTrinh in listloaict:
				if excel.objects.filter(filename=file,account=f,loaiCt=chuongTrinh).exists():
					emptyList=[]
					continue
				else:
					emptyList.append("{}".format(str(chuongTrinh)))

			if emptyList != []:
				annouceExist=annouceExist+" {}_no_Data ".format(str(f))
				continue
			else:
				annouceExist=annouceExist+" {}_Success ".format(str(f))
			title = 'THÔNG BÁO VỀ CHƯƠNG TRÌNH KHUYẾN MÃI' #.encode('utf-8')
			date_year = "Tp.HCM, Ngày {}".format(str(datetime.now().date().strftime("%d/%m/%Y")))
			pdf = PDF()
			pdf.add_font('arial','',r"static/Fonts/arial.ttf",uni=True)
			pdf.add_font('arial','B',r"static/Fonts/FontsFree-Net-arial-bold.ttf",uni=True)
			pdf.add_font('arial','I',r"static/Fonts/Arial-Italic.ttf",uni=True)
			pdf.set_font('arial','', size=10)
			# pdf=FPDF("L", "mm", "A4")

			pdf.add_page("L")
			title_w = pdf.get_string_width(title) + 6

			doc_w = pdf.w
			center = (doc_w - title_w) / 2
			#tittle
			pdf.text(center,15,title)
			#logo
			pdf.image('static/image/kimberlylogo.jpg', x = 5, w = 60,h=10, y=5,type="jpg")
			# date year

			pdf.set_font('arial','', size=8)
			pdf.text(255,20,date_year)
			# Kính gửi
			pdf.set_font('arial', 'B', size=9)
			text = 'Kính gửi: Quý Khách Hàng Kênh Hiện Đại'

			pdf.text(10, 25, text)
			# Tên chương trình
			pdf.set_font('arial','', size=8)
			text = "Công ty TNHH Kimberly-Clark Việt Nam (Công ty)  xin trân trọng thông báo chương trình đến Quý Khách Hàng như thông tin đính kèm,"

			pdf.text(10,32,text)
			# Loại CT

			#loai ct,loai acc
			print(pdf.get_x())
			print(pdf.get_y())
			pdf.set_y(pdf.get_y()+30)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"Loại CT",0,0,"L",1)

			pdf.cell(80,5,"{}".format(",".join(listloaict),0,1,"L",1))
			pdf.set_fill_color(153,204,255)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"Account","B",0,"L",1)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"{}".format(f),"B",1,"L",1)
			pdf.set_y(pdf.get_y()+5)
			#table header
			headers=['Mechanics: get/discount',"Product","Post start date","Post end date"]
			for i,header in enumerate(headers):
				if i==0:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"{}".format(headers[i]),"B",0,"L",1)
				elif i==1:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"{}".format(headers[i]),"B",0,"L",1)
				elif i ==2:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"{}".format(headers[i]),"B",0,"L",1)
				else:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"{}".format(headers[i]),"B",1,"L",1)

			#table data
			for ct in listloaict:
				listMecha=excel.objects.filter(filename=file,account=f,loaiCt=ct).values("mechanicsGetORDiscount").distinct()
				for mecha in listMecha:
					datas=excel.objects.filter(filename=file,account=f,loaiCt=ct,mechanicsGetORDiscount=mecha.get("mechanicsGetORDiscount"))
					print(datas)
					for row,data in enumerate(datas):
						for col,colAlphabet in enumerate(["A","B","C","D"]):
							
							mechanicsString=data.mechanicsGetORDiscount.replace("\n","")
							if row >0:
								mechanicsString=""	
							cellWitdhMax=80
							if pdf.get_string_width(mechanicsString) < cellWitdhMax:
								if row ==(len(datas)-1):				
									if headers[col]=='Mechanics: get/discount':
										string=mechanicsString
										pdf.cell(80,5,string,"B",0,"L")
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(80,5,"",0,0,"L")
										else:
											pdf.cell(80,5,string,0,0,"L")
									elif headers[col]=="Post start date":
										pdf.cell(60,5,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									elif headers[col]=="Post end date":
										pdf.cell(60,5,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
								else:
									if headers[col]=='Mechanics: get/discount':
										string=mechanicsString
										pdf.cell(80,5,string,0,0,"L")
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(80,5,"",0,0,"L")
										else:
											pdf.cell(80,5,string,0,0,"L")
									elif headers[col]=="Post start date":
										pdf.cell(60,5,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									elif headers[col]=="Post end date":
										pdf.cell(60,5,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
							else:
								if row ==(len(datas)-1):
									mechanicsStringLen=len(mechanicsString)
									startChar=0
									maxChar=0
									holdStringEachLine=[]
									holdStringTemp=""
									while (startChar<=mechanicsStringLen):
										while (pdf.get_string_width(holdStringTemp) < cellWitdhMax) and ((startChar+maxChar)<=mechanicsStringLen):
											maxChar=maxChar+1
											holdStringTemp= mechanicsString[startChar:(maxChar+startChar)]						
										startChar=startChar+maxChar
										holdStringEachLine.append(holdStringTemp)
										#reset
										maxChar=0
										holdStringTemp=""					
									print(holdStringEachLine)
									line=len(holdStringEachLine) #Numbers of line
									if headers[col]=='Mechanics: get/discount':
										
										xPos=pdf.get_x()
										yPos=pdf.get_y()
										pdf.multi_cell(cellWitdhMax,5,mechanicsString,"B")
										pdf.set_xy(xPos+cellWitdhMax,yPos)
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(80,5*line,"",0,0,"L")
										else:
											pdf.cell(80,5*line,string,0,0,"L",)
									elif headers[col]=="Post start date":
										pdf.cell(60,5*line,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									elif headers[col]=="Post end date":
										pdf.cell(60,5*line,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
								else:
									mechanicsStringLen=len(mechanicsString)
									startChar=0
									maxChar=0
									holdStringEachLine=[]
									holdStringTemp=""
									while (startChar<=mechanicsStringLen):
										while (pdf.get_string_width(holdStringTemp) < cellWitdhMax) and ((startChar+maxChar)<=mechanicsStringLen):
											maxChar=maxChar+1
											holdStringTemp= mechanicsString[startChar:(maxChar+startChar)]						
										startChar=startChar+maxChar
										holdStringEachLine.append(holdStringTemp)
										#reset
										maxChar=0
										holdStringTemp=""					
									print(holdStringEachLine)
									line=len(holdStringEachLine) #Numbers of line
									if headers[col]=='Mechanics: get/discount':
										
										xPos=pdf.get_x()
										yPos=pdf.get_y()
										pdf.multi_cell(cellWitdhMax,5,mechanicsString,0)
										pdf.set_xy(xPos+cellWitdhMax,yPos)
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(80,5*line,"",0,0,"L")
										else:
											pdf.cell(80,5*line,string,0,0,"L",)
									elif headers[col]=="Post start date":
										pdf.cell(60,5*line,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									elif headers[col]=="Post end date":
										pdf.cell(60,5*line,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")

			#table footer
			for i,header in enumerate(headers):
				if i==0:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"","B",0,"L",1)
				elif i==1:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"","B",0,"L",1)
				elif i ==2:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"","B",0,"L",1)
				else:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"","B",1,"L",1)
					
			#add footer
			pdf.set_y(pdf.get_y()+5)
			pdf.cell(100,5,"Mong Quý Khách Hàng cùng hợp tác với Công ty để đảm bảo các chương trình thực thi hiệu quả trong thời gian tới.",0,1)
			pdf.cell(100,5,"Nếu Quý Khách Hàng có bất kỳ vấn đề nào cần làm rõ, vui lòng cho KCV được biết để cùng trao đổi",0,1)

			pdf.set_y(pdf.get_y()+5)
			pdf.cell(40,5,"Trân trọng cảm ơn Quý Khách Hàng",0,1)
			pdf.cell(40,5,"Trưởng bộ phận quản lý kênh hiện đại",0,1)
			
			
			pdf.set_y(pdf.get_y()+5)
			pdf.cell(40,5,"Phạm Nguyên Thủ",0,1)
			
			
		
			pos="{}x{}".format(round(pdf.get_x()),round(pdf.get_y()))
			print(pos)
			maxpageIndex=pdf.page-1
			print(maxpageIndex)
			pdf.output("pdffile.pdf")
			# pdf.write_html(html)
			filename="ThưThôngBáo_{}_{}.pdf".format(f,str(datetime.now().date()))
			print(filename)
			with open("pdffile.pdf",'rb') as pdf:
				print(1)
				pdffile=pdfFile()

				pdffile.masterFile=file
				pdffile.creator=user
				pdffile.account=pdfAccount
				pdffile.loaict=",".join(listloaict)
				pdffile.pos=pos
				pdffile.page_number=maxpageIndex
				pdffile.slaveFile.save(filename,File(pdf))
				if pdfEmail:			
					for email in pdfEmail:
						print(email)
						pdffile.emailExtracted.add(email)
				pdffile.save()
	elif loaiAccount!="all" and loaict!="all":
		print("4")
		listAcc=loaiAccount.split(',')
		listCt=loaict.split(',')
		
		for f in listAcc:
			emptyList=[]
			pdfAccount=excelAccount.objects.get(account=f)
			try:
				pdfEmail=accountEmail.objects.filter(account=pdfAccount)
			except:
				pdfEmail=[]
			for chuongTrinh in listCt:
				if excel.objects.filter(filename=file,account=f,loaiCt=chuongTrinh).exists():
					emptyList=[]
					continue
				else:
					emptyList.append("{}".format(str(chuongTrinh)))

			if emptyList != []:
				annouceExist=annouceExist+" {}_no_Data ".format(str(f))
				continue
			else:
				annouceExist=annouceExist+" {}_Success ".format(str(f))
			title = 'THÔNG BÁO VỀ CHƯƠNG TRÌNH KHUYẾN MÃI' #.encode('utf-8')
			date_year = "Tp.HCM, Ngày {}".format(str(datetime.now().date().strftime("%d/%m/%Y")))
			pdf = PDF()
			pdf.add_font('arial','',r"static/Fonts/arial.ttf",uni=True)
			pdf.add_font('arial','B',r"static/Fonts/FontsFree-Net-arial-bold.ttf",uni=True)
			pdf.add_font('arial','I',r"static/Fonts/Arial-Italic.ttf",uni=True)
			pdf.set_font('arial','', size=10)
			# pdf=FPDF("L", "mm", "A4")

			pdf.add_page("L")
			title_w = pdf.get_string_width(title) + 6

			doc_w = pdf.w
			center = (doc_w - title_w) / 2
			#tittle
			pdf.text(center,15,title)
			#logo
			pdf.image('static/image/kimberlylogo.jpg', x = 5, w = 60,h=10, y=5,type="jpg")
			# date year

			pdf.set_font('arial','', size=8)
			pdf.text(255,20,date_year)
			# Kính gửi
			pdf.set_font('arial', 'B', size=9)
			text = 'Kính gửi: Quý Khách Hàng Kênh Hiện Đại'

			pdf.text(10, 25, text)
			# Tên chương trình
			pdf.set_font('arial','', size=8)
			text = "Công ty TNHH Kimberly-Clark Việt Nam (Công ty)  xin trân trọng thông báo chương trình đến Quý Khách Hàng như thông tin đính kèm,"

			pdf.text(10,32,text)
			# Loại CT

			#loai ct,loai acc
			print(pdf.get_x())
			print(pdf.get_y())
			pdf.set_y(pdf.get_y()+30)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"Loại CT",0,0,"L",1)

			pdf.cell(80,5,"{}".format(loaict),0,1,"L",1)
			pdf.set_fill_color(153,204,255)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"Account","B",0,"L",1)
			pdf.set_fill_color(153,204,255)
			pdf.cell(80,5,"{}".format(f),"B",1,"L",1)
			pdf.set_y(pdf.get_y()+5)
			#table header
			headers=['Mechanics: get/discount',"Product","Post start date","Post end date"]
			for i,header in enumerate(headers):
				if i==0:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"{}".format(headers[i]),"B",0,"L",1)
				elif i==1:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"{}".format(headers[i]),"B",0,"L",1)
				elif i ==2:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"{}".format(headers[i]),"B",0,"L",1)
				else:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"{}".format(headers[i]),"B",1,"L",1)

			#table data
			for ct in listCt:
				listMecha=excel.objects.filter(filename=file,account=f,loaiCt=ct).values("mechanicsGetORDiscount").distinct()
				for mecha in listMecha:
					datas=excel.objects.filter(filename=file,account=f,loaiCt=ct,mechanicsGetORDiscount=mecha.get("mechanicsGetORDiscount"))
					print(datas)
					for row,data in enumerate(datas):
						for col,colAlphabet in enumerate(["A","B","C","D"]):
							
							mechanicsString=data.mechanicsGetORDiscount.replace("\n","")
							if row >0:
								mechanicsString=""	
							cellWitdhMax=80
							if pdf.get_string_width(mechanicsString) < cellWitdhMax:
								if row ==(len(datas)-1):				
									if headers[col]=='Mechanics: get/discount':
										string=mechanicsString
										pdf.cell(80,5,string,"B",0,"L")
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(80,5,"",0,0,"L")
										else:
											pdf.cell(80,5,string,0,0,"L")
									elif headers[col]=="Post start date":
										pdf.cell(60,5,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									elif headers[col]=="Post end date":
										pdf.cell(60,5,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
								else:
									if headers[col]=='Mechanics: get/discount':
										string=mechanicsString
										pdf.cell(80,5,string,0,0,"L")
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(80,5,"",0,0,"L")
										else:
											pdf.cell(80,5,string,0,0,"L")
									elif headers[col]=="Post start date":
										pdf.cell(60,5,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									elif headers[col]=="Post end date":
										pdf.cell(60,5,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
							else:
								if row ==(len(datas)-1):
									mechanicsStringLen=len(mechanicsString)
									startChar=0
									maxChar=0
									holdStringEachLine=[]
									holdStringTemp=""
									while (startChar<=mechanicsStringLen):
										while (pdf.get_string_width(holdStringTemp) < cellWitdhMax) and ((startChar+maxChar)<=mechanicsStringLen):
											maxChar=maxChar+1
											holdStringTemp= mechanicsString[startChar:(maxChar+startChar)]						
										startChar=startChar+maxChar
										holdStringEachLine.append(holdStringTemp)
										#reset
										maxChar=0
										holdStringTemp=""					
									print(holdStringEachLine)
									line=len(holdStringEachLine) #Numbers of line
									if headers[col]=='Mechanics: get/discount':
										
										xPos=pdf.get_x()
										yPos=pdf.get_y()
										pdf.multi_cell(cellWitdhMax,5,mechanicsString,"B")
										pdf.set_xy(xPos+cellWitdhMax,yPos)
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(80,5*line,"",0,0,"L")
										else:
											pdf.cell(80,5*line,string,0,0,"L",)
									elif headers[col]=="Post start date":
										pdf.cell(60,5*line,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									elif headers[col]=="Post end date":
										pdf.cell(60,5*line,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")
								else:
									mechanicsStringLen=len(mechanicsString)
									startChar=0
									maxChar=0
									holdStringEachLine=[]
									holdStringTemp=""
									while (startChar<=mechanicsStringLen):
										while (pdf.get_string_width(holdStringTemp) < cellWitdhMax) and ((startChar+maxChar)<=mechanicsStringLen):
											maxChar=maxChar+1
											holdStringTemp= mechanicsString[startChar:(maxChar+startChar)]						
										startChar=startChar+maxChar
										holdStringEachLine.append(holdStringTemp)
										#reset
										maxChar=0
										holdStringTemp=""					
									print(holdStringEachLine)
									line=len(holdStringEachLine) #Numbers of line
									if headers[col]=='Mechanics: get/discount':
										
										xPos=pdf.get_x()
										yPos=pdf.get_y()
										pdf.multi_cell(cellWitdhMax,5,mechanicsString,0)
										pdf.set_xy(xPos+cellWitdhMax,yPos)
									if headers[col]=='Product':
										string=data.product
										if string== None:
											pdf.cell(80,5*line,"",0,0,"L")
										else:
											pdf.cell(80,5*line,string,0,0,"L",)
									elif headers[col]=="Post start date":
										pdf.cell(60,5*line,"{}".format(data.postStartDate.strftime("%d/%m/%Y")),0,0,"L")
									elif headers[col]=="Post end date":
										pdf.cell(60,5*line,"{}".format(data.postEndDate.strftime("%d/%m/%Y")),0,1,"L")

			#table footer
			for i,header in enumerate(headers):
				if i==0:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"","B",0,"L",1)
				elif i==1:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(80,5,"","B",0,"L",1)
				elif i ==2:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"","B",0,"L",1)
				else:
					pdf.set_font('arial', 'B', size=9)
					pdf.set_fill_color(153,204,255)
					pdf.cell(60,5,"","B",1,"L",1)
					
			#add footer
			pdf.set_y(pdf.get_y()+5)
			pdf.cell(100,5,"Mong Quý Khách Hàng cùng hợp tác với Công ty để đảm bảo các chương trình thực thi hiệu quả trong thời gian tới.",0,1)
			pdf.cell(100,5,"Nếu Quý Khách Hàng có bất kỳ vấn đề nào cần làm rõ, vui lòng cho KCV được biết để cùng trao đổi",0,1)

			pdf.set_y(pdf.get_y()+5)
			pdf.cell(40,5,"Trân trọng cảm ơn Quý Khách Hàng",0,1)
			pdf.cell(40,5,"Trưởng bộ phận quản lý kênh hiện đại",0,1)
			
			
			pdf.set_y(pdf.get_y()+5)
			pdf.cell(40,5,"Phạm Nguyên Thủ",0,1)
			
			
		
			pos="{}x{}".format(round(pdf.get_x()),round(pdf.get_y()))
			print(pos)
			maxpageIndex=pdf.page-1
			print(maxpageIndex)
			pdf.output("pdffile.pdf")
			# pdf.write_html(html)
			filename="ThưThôngBáo_{}_{}.pdf".format(f,str(datetime.now().date()))
			print(filename)
			with open("pdffile.pdf",'rb') as pdf:
				print(1)
				pdffile=pdfFile()

				pdffile.masterFile=file
				pdffile.creator=user
				pdffile.account=pdfAccount
				pdffile.loaict=loaict
				pdffile.pos=pos
				pdffile.page_number=maxpageIndex
				pdffile.slaveFile.save(filename,File(pdf))
				if pdfEmail:			
					for email in pdfEmail:
						print(email)
						pdffile.emailExtracted.add(email)
				pdffile.save()
	return annouceExist


