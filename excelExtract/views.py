from django.shortcuts import render,redirect
import sys
from rest_framework.decorators import api_view
from excelExtract.forms import uploadDocumentForm
from excelExtract.models import accountEmail, document, excelAccount,pdfFile, excel
from . import excelExtract
from user.models import Profile, User
from django.db import transaction
from rest_framework import status
from django.conf import settings
from django.http import HttpResponse
from rest_framework.response import Response
from django.views.decorators.clickjacking import xframe_options_sameorigin
from utils import send_email
from datetime import datetime
import requests
import json
from django.core.files import File
from django.db.models import Q
import fitz
import openpyxl
from openpyxl.styles import Alignment,NamedStyle,PatternFill,Font
from openpyxl.writer.excel import save_virtual_workbook
from user.forms import LoginForm
import os
import zipfile
import io
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
import logging
from itertools import chain
import re
logger = logging.getLogger("debug_purposes")
#Function to find the position of texte, and then return the coordinate of its to insert the signature
def detect_position(pdf_file_location):
	pdf = fitz.open(pdf_file_location)
	page_0 = pdf.load_page(0)
	page_width, page_height = page_0.rect.width, page_0.rect.height
	y1 = 200
	marked_page_num=0
	x0 = 200
	search_text = 'Trưởng bộ phận quản lý kênh hiện đại'
	for i in range(pdf.page_count):
		text_instances=pdf.load_page(i).search_for(search_text)
		if  text_instances != []:
			marked_page_num = i
			x0=text_instances[0].x0
			y1=text_instances[0].y1			
	if y1 >  595:
		marked_page_num = marked_page_num + 1
		left = x0
		bottom = 10.5 * 3
	else:
		left = x0
		bottom = y1 
	# convert to milimeter  
	left = left
	bottom = (page_height - bottom) - 10.5*2 - 50  # 10.5 la chieu cao 1 dong dong, 50 trong 50x50
	return left, bottom, marked_page_num

#DJANGO VIEWS FUNCTIONS
#----------------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------
@xframe_options_sameorigin
def kcToolPage(request):
	if request.user.is_authenticated:
		form=uploadDocumentForm()
		files=document.objects.all().order_by("-id")
		num_files=len(files)
		if request.method=='POST':
			# with transaction.atomic():
			# 	try:
					file=request.FILES['Excel_File']
					print(file)
					if file:
						try:
							user=request.user
							profile=Profile.objects.get(user=user)
							excelExtract.importDataExcel(file, user=profile)
						except:
							excelExtract.importDataExcel(file)
				# except:
				# 	pass
		return render(request,"index.html",{"form":form,"files":files,"num_files":num_files, "active_id":1})
	else :
		form = LoginForm()
		return render(request, 'login.html', {'form':form})


def newCreatedDocs(request):
	if request.user.is_authenticated:
		user=request.user
		unconfirmpdfs=pdfFile.objects.filter(confirmed=False,is_deleted=False).order_by("-id")
		numberunconfirmpdfs=0
		if request.user.is_admin:
			for pdf in pdfFile.objects.filter(confirmed=False,is_deleted=False):
				if pdf.account in excelAccount.objects.filter(responsibleBy__isnull=False):
					numberunconfirmpdfs+=1
		else:
			user=request.user
			profile=Profile.objects.get(user=user)
			
			for pdf in pdfFile.objects.filter(confirmed=False,is_deleted=False):
				if pdf.account in excelAccount.objects.filter(responsibleBy=profile).all():
					numberunconfirmpdfs+=1
		accountList=excelAccount.objects.filter(responsibleBy__isnull=False)
		print(accountList)
		profile=Profile.objects.get(user=user)
		listaccount=excelAccount.objects.filter(responsibleBy=profile)
		if request.GET.get("account",None):
			account=excelAccount.objects.filter(account=request.GET.get("account"))[0]
			print(account.pk)
			unconfirmpdfs = pdfFile.objects.filter(confirmed=False,account=account,is_deleted=False).order_by("-id")
		return render(request,"KCtool/vb-cho-xac-nhan.html",{"numberunconfirmpdfs":numberunconfirmpdfs,"unconfirmpdfs":unconfirmpdfs, "URL":settings.URL, "active_id":2,"accountList":accountList,"user":user,"listaccount":listaccount,"account":request.GET.get("account",None)})
	else :
		form = LoginForm()
		return render(request, 'login.html', {'form':form}) 

def confirmedDocs(request):
	if request.user.is_authenticated:
		user=request.user
		numberconfirmedpdfs=len(pdfFile.objects.filter(confirmed=True,sended=False,signed=False))
		pdfs=pdfFile.objects.filter(confirmed=True,sended=False,signed=False).order_by("-id")
		accountList=excelAccount.objects.all()	
		return render(request,"KCtool/vb-da-xac-nhan.html",{"numberconfirmedpdfs":numberconfirmedpdfs,"pdfs":pdfs, "active_id":3,"user":user,"accountList":accountList,"key_word":request.GET.get("key_word",""),"account":request.GET.get("account",None),"fromdate":request.GET.get("fromdate"),"todate":request.GET.get("todate"),"fromdate2":request.GET.get("fromdate2"),"todate2":request.GET.get("todate2")})

	else :
		form = LoginForm()
		return render(request, 'login.html', {'form':form})

def signedDocs(request):
	if request.user.is_authenticated:
		user=request.user
		numbersendedpdfs=len(pdfFile.objects.filter(sended=True))
		pdfs=pdfFile.objects.filter(confirmed=True,signed=True,sended=True).order_by("-id")
		accountList=excelAccount.objects.all()
		set_loai_ct=set()
		for pdf in pdfs:
			loaict=(pdf.loaict).split(',')
			for ct in loaict:
				set_loai_ct.add(ct)
		list_loai_ct=list(set_loai_ct)
		number_qr=0
		if request.GET.getlist("list_ac") or request.GET.getlist("list_ct") or request.GET.get('daterange') :
			first_qs=pdfFile.objects.none()
			if  request.GET.getlist("list_ac"):
				if  'All' in request.GET.getlist("list_ac"):
					first_qs=pdfFile.objects.filter(confirmed=True,signed=True,sended=True).order_by("-id")
				else:
					for f in request.GET.getlist("list_ac"):
						account=excelAccount.objects.get(account=f)
						current_qs=pdfFile.objects.filter(confirmed=True,signed=True,sended=True,account=account).order_by("-id")
						first_qs=first_qs|current_qs 
			else:
				first_qs=pdfFile.objects.filter(confirmed=True,signed=True,sended=True).order_by("-id")
			if request.GET.getlist("list_ct"):
				second_qs=pdfFile.objects.none()
				if  'All' in request.GET.getlist("list_ct"):
					second_qs=first_qs
				else:
					for f in request.GET.getlist("list_ct"):
						print(first_qs.filter(loaict__icontains=f))
						second_qs=second_qs|first_qs.filter(loaict__icontains=f)
			else:
				second_qs=first_qs
			if request.GET.get('daterange') :
				daterange=request.GET.get('daterange').split(' - ')
				start_date=datetime.strptime(daterange[0], '%m/%d/%Y')
				end_date=datetime.strptime(daterange[1], '%m/%d/%Y')
				third_qs=second_qs.filter(SignedTime__gte=start_date,SignedTime__lte=end_date)
			else:
				third_qs=second_qs
			number_qr=len(third_qs)
			pdfs=third_qs
		if request.GET.get("fromdate2",None) and request.GET.get("fromdate2",None):	
			fromdate2=request.GET.get("fromdate2")
			todate2=request.GET.get("todate2")
			print(fromdate2,type(fromdate2))
			# actlogs = pdfFile.objects.filter(createdTime__date__gte=fromdate2,createdTime__date__lte=todate2)|pdfFile.objects.filter(sendingTime__date__gte=fromdate2,createdTime__date__lte=todate2)
			return export_virtual_excel(fromdate2,todate2)
			
			
		return render(request,"KCtool/vb-da-ky.html",{"numbersendedpdfs":numbersendedpdfs,"pdfs":pdfs, "active_id":4,"user":user,"accountList":accountList,"fromdate2":request.GET.get("fromdate2"),"todate2":request.GET.get("todate2"),"list_loai_ct":list_loai_ct,'number_qr':number_qr})

	else :
		form = LoginForm()
		return render(request, 'login.html', {'form':form})

def untrackedDocs(request):
	if request.user.is_authenticated:
		untrackedAccount=excelAccount.objects.filter(responsibleBy__isnull=True)
		Alldocs=pdfFile.objects.filter(is_deleted=False).order_by("-createdTime")
		return render(request,"KCtool/vb-chua-duoc-quan-ly.html",{"Alldocs":Alldocs,"untrackedAccount":untrackedAccount, "active_id":5})

	else :
		form = LoginForm()
		return render(request, 'login.html', {'form':form})
def staffManager(request):
	if request.user.is_authenticated:
		if request.user.is_admin :
			staffList=Profile.objects.all().order_by("-pk")
			numberstaff=len(staffList)
			list_account=excelAccount.objects.all().order_by('account')
			pages=Paginator(list_account,10)	
			if pages.num_pages > 2:
				for index in range(1,100):
					if Paginator(list_account,index).num_pages == 2:
						pages=Paginator(list_account,index)
						break
			# for page in pages:
			# 	for account in page:
			# 		print(page,account)
			return render(request,"KCtool/quan-ly-nhan-su.html",{"staffList":staffList,"numberstaff":numberstaff, "active_id":6,'pages':pages})
	else :
		form = LoginForm()
		return render(request, 'login.html', {'form':form})





@api_view(["POST"])
def addNewProfile(request):
	data=request.data
	print(data)
	if not data["full_name_create"]:
		return Response({"full_name_required":"Full name required"}, status=status.HTTP_200_OK)
	if not data['new_email_create']:
		return Response({"Email_err":"Email required"}, status=status.HTTP_200_OK)
	else:
		if Profile.objects.filter(email=data['new_email_create']).exists():
			return Response({"Email_err":"Email existed"}, status=status.HTTP_200_OK)
		else:
			if not re.fullmatch(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b' ,data['new_email_create']):
				return Response({"Email_err":"Invalid Email"}, status=status.HTTP_200_OK)
					
	if not (data['user_name_create']):
		return Response({"user_input_err":"Please enter Username"}, status=status.HTTP_200_OK)
	else:
		if User.objects.filter(user_name=data['user_name_create']).exists():
			return Response({"user_input_err":"Username existed"}, status=status.HTTP_200_OK)
	if data['password_create']:
		if not re.fullmatch(r'[A-Za-z0-9@#$%^&+=]{6,}', data['password_create']):
			return Response({"password_err":"Invalid password"}, status=status.HTTP_200_OK)
	else:
		return Response({"password_err":"Password required"}, status=status.HTTP_200_OK)
	with transaction.atomic():
		try:
			user_name=data['user_name_create']
			password=data['password_create']
			user=User.objects.create(user_name=user_name)
			user.set_password(password)
			if data['admin_role']:
				user.is_admin=True
			if data['upload_role']:
				user.is_uploader=True
			if data['sign_role']:
				user.is_signer=True
			user.save()
			userProfile=Profile.objects.create(user=user)
			try:
				userProfile.phone_number=data["phone_number_create"]
			except:
				userProfile.phone_number=None
			try:
				userProfile.email=data["new_email_create"]
			except:
				raise Exception()
			userProfile.full_name=data["full_name_create"]
			userProfile.save()
		except:
			return Response({"msg":"failed"}, status=status.HTTP_403_FORBIDDEN)
	return Response({"msg":"success"}, status=status.HTTP_200_OK)
def deletedDocs(request):
	deletedDocs=pdfFile.objects.filter(is_deleted=True).order_by('-pk')
	numberdeletedpdfs=len(deletedDocs)
	return render(request,"KCtool/vb-da-xoa.html",{"deletedDocs":deletedDocs,"active_id":8,"numberdeletedpdfs":numberdeletedpdfs})



def info_page(request):
	return render(request,"KCtool/thong-tin-ho-tro.html")
#----------------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------------------
def delete_profile(request):
	userId=request.GET.get("pk",None)
	try:
		ID= int(userId)
		user=User.objects.get(pk= ID)	
		user.delete()
	except:
		pass
	return redirect("KCTool:staffManager")
#DJANGO API VIEWS
#----------------------------------------------------------------------------------------------------
#CREATE PDF API
@api_view(["POST"])
def create_pdf(request):
	try:
		with transaction.atomic():
			print(request.data)
			id_excel = int(request.data["pk_excel"])
			values_category = request.data["values_category"]
			values_account = request.data["values_account"]
			print(values_account)
			print(values_category)
			user=request.user
			profile = Profile.objects.get(user=user)
			annouce=excelExtract.exportFiles(loaict=values_category,fileID=id_excel,loaiAccount=values_account,user=profile) 
			if values_account != "all":
				listAccount=values_account.split(",")
				listEmail=[]
				for account in listAccount:
					try:
						email=excelAccount.objects.get(account=account).responsibleBy
						
						if email.email :
							if email.email not in listEmail:
								listEmail.append(email.email)
					except:
						pass
				send_email.send_noti_to_confirmer(listEmail)
				print(listEmail)
			else:
				
				file=document.objects.get(pk=id_excel)
				listEmail=[]
				listAccount=[]
				
				for f in excel.objects.filter(filename=file):
					if f.account not in listAccount:
						listAccount.append(f.account)
				
				for account in listAccount:
					try:
						email=excelAccount.objects.get(account=account)
						email=email.responsibleBy			
						if email.email :
							if email.email not in listEmail:
								listEmail.append(email.email)
					except:
						pass
				send_email.send_noti_to_confirmer(listEmail)			
				
			return Response({"annouce":annouce}, status=status.HTTP_200_OK)
	except:
		return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#API TO RETURN DYNAMIC OPTIONS OF AN EXCEL FILE INCLUDE LOAI CHUONG TRINH,LOAI ACCOUNT
@api_view(["POST"])
def getListAccount(request): 
	# try:
		arr_listaccounts=[]
		arr_loaiCt = []
		list_accounts = ""
		list_loaiCt = ""
		pk = request.data["pk_excel"]
		file=document.objects.get(pk=int(pk))
		list_accounts +=f'''
					<option group="list_accounts_group" value="all" data-badge="">All</option>
				'''
		list_loaiCt+= f'''
					<option group="list_LoaiCT_group" value="all" data-badge="">All</option>
				'''
		for f in excel.objects.filter(filename=file):
			if f.account not in arr_listaccounts:

				arr_listaccounts.append(f.account)
				list_accounts += f'''
					<option group="list_accounts_group" value="{f.account}" data-badge="">{f.account}</option>
				'''
			if f.loaiCt not in arr_loaiCt:
				arr_loaiCt.append(f.loaiCt)
				list_loaiCt += f'''
					<option group="list_LoaiCT_group" value="{f.loaiCt}" data-badge="">{f.loaiCt}</option>
				'''
		
		return Response({"Status":"Success", "list_accounts": list_accounts, "list_loaiCt":list_loaiCt},  status=status.HTTP_200_OK)
	# except:
	# 	return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)
#SEND NOTIFY TO THE SIGNER
@api_view(["POST"])
def confirm_pdf(request): 
	try:
		with transaction.atomic():
			user = request.user
			profile = Profile.objects.get(user=user)
			list_id_pdf_file = request.data["list_id_pdf_file"]
			list_id=list_id_pdf_file.split(",")
			for i in list_id:
				pdf = pdfFile.objects.get(pk=int(i))
				pdf.confirmed = True
				pdf.confirmer= profile
				pdf.confirmedTime= datetime.now()
				pdf.save()
			print(list_id_pdf_file)
			print(list_id)
			numberunsignepdfs=len(pdfFile.objects.filter(signed=False,sended=False,confirmed=True))
			signer=User.objects.filter(is_signer=True)[0]
			signerName=str(Profile.objects.filter(user=signer)[0].email)
			print(signerName)
			send_email.send_noti_to_partner_sign_by_email2([], signerName)
			return Response({"code":"00"}, status=status.HTTP_200_OK)
	except:
		err_mess = sys.exc_info()[0].__name__ + ": "+ str(sys.exc_info()[1])
		print(err_mess)
		return Response({"err":err_mess},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
#SIGN AND SEND PDF
@api_view(["POST"])
def sign_and_send_pdf(request): 
	account_data={"user_name": "thach.nguyenphamngoc@kcc.com",
  				"password": "PVS@@123456"}
	headers = { 
	'Content-Type':'application/json' 
   }
	response_obj = requests.post(r"https://api.pvs.com.vn/user-api/api/token/", data=json.dumps(account_data),headers=headers)
	token=response_obj.json()['data']['access']
	print(token)
	print("response_obj.status_code",response_obj.status_code)
	if response_obj.status_code >= 200 and response_obj.status_code<300:
		try:
			
				log=""
				list_id_pdf_file = request.data["list_id_pdf_file"]
				
				list_id=list_id_pdf_file.split(",")
				accountCate=[]
				
				for i in list_id:
					account = str(pdfFile.objects.get(pk=int(i)).account)
					if account not in accountCate:
						accountCate.append(account)
				print(accountCate)
				for account in accountCate:
					listfile=[]
					listct=[]
					for i in list_id:
						if account == str(pdfFile.objects.get(pk=int(i)).account):
							print(account)
							pdf=pdfFile.objects.get(pk=int(i))	
							pdffile=pdfFile.objects.get(pk=int(i)).slaveFile
							linkfile=settings.URL+"/"+str(pdffile)
							chuongtrinh=str(pdfFile.objects.get(pk=int(i)).loaict).split(",")
							print("chuongtrỉnh",chuongtrinh)
							for ct in chuongtrinh:
								if ct not in listct:
									listct.append(ct)
							tple=detect_position(str(pdffile))
							
							print(tple)
							data_send ={
								"pdf_url":linkfile,
								"sign_pos": "{}x{}".format(round(tple[0]),round(tple[1])),
								"contact": "thach.nguyenphamngoc@kcc.com",
								"reason": "sign contract",
								"page_number":tple[2]
							}							
							headers2 = { 'Content-Type':'application/json', 
							'Authorization': 'Bearer ' + token }
							print(data_send)
							response_obj2 = requests.post(r"https://api.pvs.com.vn/e-invoice-api/api/ca-sign/sign-pdf/58", data=json.dumps(data_send), headers=headers2)				
							if response_obj2.status_code  >= 200 and response_obj2.status_code<300:
								log+="{}:success ".format(str(pdffile).replace("documents/slavefiles/",""))
								binarytext=response_obj2.content
								with open("file.pdf","wb") as file:
									file.write(binarytext)
								with open ("file.pdf","rb") as file:
									name="THUTHONGBAO_{}_{}.pdf".format(str(account).upper(),str(datetime.now().date()))
									newpdf=pdf.slaveFile.save(name,File(file))
									pdf.sended=True
									pdf.signed=True
									pdf.SignedTime=datetime.now()
									pdf.sendingTime=datetime.now()
									pdf.save()
									fileurl= settings.URL+"/"+str(pdf.slaveFile)
								listfile.append(fileurl)
							if response_obj2.status_code  >= 300 and response_obj2.status_code <= 500:
								try:
									log+=str(response_obj2.json()['message'])
								except:
									log+=str(response_obj2.json()['message'])
									continue
					if listfile != []:
						listemail=[]
						for email in pdf.emailExtracted.all():
							if email not in listemail:
								listemail.append(email)
							print(listfile)
							print(email)
						send_email.send_noti_to_partner_sign_by_email(",".join(listct),account,listfile,listemail)
					else:
						log+=("{}:failed".format(account))
				return Response({"log":log}, status=status.HTTP_200_OK)
		except:
			err_mess = sys.exc_info()[0].__name__ + ": "+ str(sys.exc_info()[1])
			print(err_mess)
			return Response({"err":err_mess},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
	else :
		return Response(response_obj.json()['message'], status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#ONLY SEND PDF 
@api_view(["POST"])
def send_pdf(request): 
	try:
		log="success"
		list_id_pdf_file = request.data["list_id_pdf_file"]
		
		list_id=list_id_pdf_file.split(",")
		accountCate=[]
	
		for i in list_id:
			account = str(pdfFile.objects.get(pk=int(i)).account)
			if account not in accountCate:
				accountCate.append(account)
		for account in accountCate:
			listfile=[]
			listct=[]
			for i in list_id:
				if account == str(pdfFile.objects.get(pk=int(i)).account):
					pdf=pdfFile.objects.get(pk=int(i))	
					pdffile=pdfFile.objects.get(pk=int(i)).slaveFile
					pdf.sendingTime=datetime.now()
					pdf.save()
					print(datetime.now())
					linkfile=settings.URL+"/"+str(pdffile)
					chuongtrinh=str(pdfFile.objects.get(pk=int(i)).loaict).split(",")
					for ct in chuongtrinh:
						if ct not in listct:
							listct.append(ct)
					listfile.append(linkfile)

			if listfile != []:
				listemail=[]
				for email in pdf.emailExtracted.all():
					if email not in listemail:
						listemail.append(email)
						print(listfile)
						print(email)
				send_email.send_noti_to_partner_sign_by_email(",".join(listct),account,listfile,listemail)
			else:
				log+=("there are no files")
		return Response({"log":log}, status=status.HTTP_200_OK)
	except:
		return Response({"log":"failed"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
#DELETE THE PDF FILE
@api_view(["POST"])
def deleteFile(request): 
	print(request.data)
	if request.data["list_id_pdf_file"]:
		list_id_pdf_file = request.data["list_id_pdf_file"]
		list_id=list_id_pdf_file.split(",")
		print(list_id)
		for id in list_id:
			fileWillBeDel=pdfFile.objects.get(pk=int(id))
			fileWillBeDel.is_deleted=True
			fileWillBeDel.deletedTime=datetime.now()
			fileWillBeDel.save()
	return Response({"msg":"delete success"}, status=status.HTTP_200_OK)
#DELETE THE EXCEL FILE
@api_view(["POST"])
def deleteExcelFile(request): 
	print(request.data)
	if request.data["list_id_excel_file"]:
		list_id_pdf_file = request.data["list_id_excel_file"]
		list_id=list_id_pdf_file.split(",")
		print(list_id)
		for id in list_id:
			fileWillBeDel=document.objects.get(pk=int(id))
			print(fileWillBeDel)
			fileWillBeDel.delete()
	return Response({"msg":"delete success"}, status=status.HTTP_200_OK)
#Restore THE PDF FILE
@api_view(["POST"])
def restoreFile(request): 
	print(request.data)
	if request.data["list_id_pdf_file"]:
		list_id_pdf_file = request.data["list_id_pdf_file"]
		list_id=list_id_pdf_file.split(",")
		print(list_id)
		for id in list_id:
			fileWillBeDel=pdfFile.objects.get(pk=int(id))
			fileWillBeDel.is_deleted=False
			fileWillBeDel.save()
	return Response({"msg":"restore success"}, status=status.HTTP_200_OK)
@api_view(['POST'])
def downloadFiles(request):
	print(request.data)
	list_id_pdf_file = request.data["list_id_pdf_file"]
	list_id=list_id_pdf_file.split(",")
	filenames=[pdfFile.objects.get(pk=int(id)).slaveFile for id in list_id ]
	temp_file = ContentFile(b"", name="{}.zip".format(str(datetime.now())))
	with zipfile.ZipFile(temp_file, mode='w', compression=zipfile.ZIP_DEFLATED) as zip_file:
		files = filenames
		for file_ in files:
			path = file_.name.replace("documents/slavefiles/","")
			print(path)
			zip_file.writestr(path, file_.read())
	file_size = temp_file.tell()
	temp_file.seek(0) 
	resp = HttpResponse(temp_file, content_type='application/zip')
	resp['Content-Disposition'] = 'attachment; filename=%s' % "{}.zip".format(str(datetime.now()))
	resp['Content-Length'] = file_size
	return resp
#NOT API BUT A FUNCTION WHICH GENERATE A VIRTUAL EXCEL REPORT
def export_virtual_excel(from_date, to_date):
	col_names = ["","Account","Category","File","CreatedTime","ConfirmedTime","SendedTime","Creator","Confirmer","Sender","Confirmed","Signed","Sended"]
	actlogs = pdfFile.objects.filter(SignedTime__date__gte=from_date,SignedTime__date__lte=to_date,signed=True)

	wb = openpyxl.Workbook()
	wb.iso_dates = True
	ws = wb['Sheet']
	ws.title='Report'


	# create title
	light_blue_fill = PatternFill(start_color='87CEFA',
                   end_color='87CEFA',
                   fill_type='solid')
	font = Font(name='Arial', size=12, bold=False,
              	vertAlign=None, underline='none', strike=False,
                color='FF000000')
	for col in range(1, len(col_names)):
		_ = ws.cell(column=col,row=1,value=col_names[col])
		_.fill = light_blue_fill
		_.font = font
	alignment = Alignment(horizontal='general')
	normal_format = NamedStyle(name="normal",alignment=alignment)
	datetime_format = NamedStyle(name="datetime",number_format="DD/MMM/YYYY h:mm",alignment=alignment)
	date_format = NamedStyle(name="date",number_format="DD/MMM/YYYY",alignment=alignment)
	ws.column_dimensions['A'].width = 20
	ws.column_dimensions['B'].width = 40
	ws.column_dimensions['C'].width = 40
	ws.column_dimensions['D'].width = 20
	ws.column_dimensions['E'].width = 20
	ws.column_dimensions['F'].width = 20
	ws.column_dimensions['G'].width = 20
	ws.column_dimensions['H'].width = 20
	ws.column_dimensions['l'].width = 20

	# create content
	# now = pytz.utc.localize(datetime.utcnow())
	# now = now.replace(tzinfo=None)
	# print(now)
	for row, ticket in enumerate(actlogs, start=2):
		for col in range(1,13):		
			if col_names[col] == "Account" :
				if ticket.account:
					c = ws.cell(column=col,row=row,value=str(ticket.account))
			if col_names[col] == "Category" :
				if ticket.loaict:
					c = ws.cell(column=col,row=row,value=str(ticket.loaict))
			if col_names[col] == "file" :
				if ticket.slaveFile:
					c = ws.cell(column=col,row=row,value=str(ticket.slaveFile).replace("documents/slavefiles/","") )
			if col_names[col] == "file" :
				if ticket.slaveFile:
					c = ws.cell(column=col,row=row,value=str(ticket.slaveFile).replace("documents/slavefiles/","") )
				else:
					c = ws.cell(column=col,row=row,value="")
				c.style = datetime_format
				
			elif col_names[col] == "creator":
				c = ws.cell(column=col,row=row,value=str(ticket.creator))
				c.style = date_format
			elif col_names[col] == "CreatedTime":
				if ticket.createdTime:
					c = ws.cell(column=col,row=row,value=str(ticket.createdTime.date()))
				else:
					c = ws.cell(column=col,row=row,value="")
				c.style = normal_format
			elif col_names[col] == "ConfirmedTime":
				if ticket.confirmedTime:
					c = ws.cell(column=col,row=row,value=str(ticket.confirmedTime.date()))
				else:
					c = ws.cell(column=col,row=row,value="")
				c.style = normal_format
			elif col_names[col] == "SendedTime":
				if ticket.sendingTime:
					c = ws.cell(column=col,row=row,value=str(ticket.sendingTime.date()))
					c.style = normal_format
				else:
					c = ws.cell(column=col,row=row,value="")
			elif col_names[col] == "creator":
				if ticket.sendingTime:
					c = ws.cell(column=col,row=row,value=str(ticket.creator))
					c.style = normal_format
			elif col_names[col] == "confirmer":
				if ticket.confirmer:
					c = ws.cell(column=col,row=row,value=str(ticket.confirmer))
					c.style = normal_format
			elif col_names[col] == "sender":
				if ticket.signer:
					c = ws.cell(column=col,row=row,value=str(ticket.signer))
					c.style = normal_format		
			elif col_names[col] == "sended":
				if ticket.sended:
					c = ws.cell(column=col,row=row,value=str(ticket.sended))
					c.style = normal_format		
			elif col_names[col] == "Sended":
				if ticket.sended:
					c = ws.cell(column=col,row=row,value=str(ticket.sended))
					c.style = normal_format
			elif col_names[col] == "Signed":
				if ticket.signed:
					c = ws.cell(column=col,row=row,value=str(ticket.signed))
					c.style = normal_format	
			elif col_names[col] == "Confirmed":
				if ticket.confirmed:
					c = ws.cell(column=col,row=row,value=str(ticket.confirmed))
					c.style = normal_format		
	

	response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment;filename={}_{}.xlsx'.format(from_date,to_date)
	print(response)
	return response

@api_view(["POST"])
def updateProfile(request): ##manage staff
	with transaction.atomic():
		try:
			data=request.data
			print(data)
			staffId=int(data['staff_pk'])
			staff=Profile.objects.get(pk=staffId)
			preAccountLists= excelAccount.objects.filter(responsibleBy=staff).distinct()
			list_preAccountLists=[str(f.account) for  f in preAccountLists]
			print(list_preAccountLists)
			for account in preAccountLists: #xử lí các account bỏ tick
				if str(account) not in data['account_list_array'].split(","):
					account.responsibleBy=None
					account.save()
			for account in data['account_list_array'].split(","):
				if account in preAccountLists: # kiểm tra xem account có phải là account cũ trong db hay khôn	
					pass
				else:  #đối với account mới trong request
					try:
						acc=excelAccount.objects.get(account=account)
						acc.responsibleBy=staff
						acc.save()
					except:
						pass
				authenList=data['authen_list_array'].split(",")
				user=staff.user
				print("authenlist",authenList)
				if "is_uploader" in authenList:
					user.is_uploader =True
				else:
					user.is_uploader =False
				if "is_signer" in authenList:
					user.is_signer =True
				else:
					user.is_signer =False
				if "is_admin" in authenList:
					user.is_admin =True
				else:
					user.is_admin =False
				user.save()
				staff.full_name = data['full_name']
				staff.phone_number = data['phone']
				staff.email = data['email']
				staff.save()
				return Response({"msg":"success"}, status=status.HTTP_200_OK)
		except:
			raise Exception()
	


#UPDATE USER PROFILE
@api_view(["POST"])
def update_profile(request):
	user = request.user
	profile = Profile.objects.get(user=user)
	profile.full_name = request.data['full_name']
	profile.phone_number = request.data['phone_number']
	profile.email = request.data['email']
	profile.company_name = request.data['company_name']
	profile.position= request.data['position']
	profile.address = request.data['address']
	profile.save()
	return Response({"msg":"success"}, status=status.HTTP_200_OK)
@api_view(["POST"])
def get_info(request):
	pk=request.data['staff_pk']
	staff=Profile.objects.get(pk=pk)
	user_of_staff=staff.user
	list_responsibled_account=excelAccount.objects.filter(responsibleBy=staff).values('account')
	return Response({ 'msg':'success','full_name':staff.full_name,'phone_number':staff.phone_number,'email':staff.email,'user_name':user_of_staff.user_name,'is_admin':user_of_staff.is_admin,'is_uploader':user_of_staff.is_uploader ,'is_signer':user_of_staff.is_signer,'list_responsibled_account':list_responsibled_account}, status=status.HTTP_200_OK)